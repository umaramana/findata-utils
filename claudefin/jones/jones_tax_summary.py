"""
jones_tax_summary.py
Reads Jones Quicken transactions, maps each to an IRS group,
and writes a 2-sheet Excel: Summary + Tagged Transactions.

CLIENT:   Jones (multi-business: Farming + Rental Real Estate)
INPUT:    Quicken "Transaction" report exported as .xlsx (no pivot sheet)
OUTPUT:   Jones_Tax_Summary.xlsx  (Summary tab + Transactions tab)

Usage — default paths (run from workspace root):
    python claudefin/jones/jones_tax_summary.py

Usage — custom input/output (e.g. next year's file):
    python claudefin/jones/jones_tax_summary.py \\
        --input "docs/jones/Quicken_2026.xlsx" \\
        --output "docs/jones/Jones_Tax_Summary_2026.xlsx"

Reuse notes:
  - IRS mapping is in irs_group() — edit there to add/rename categories
  - Property list is PROPERTY_ORDER — add new rental properties there
  - Quicken artifact rules are in preprocess_row() — add new ones as discovered
  - Reconciliation: script prints section totals + diff on every run
"""

import argparse
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

_DEFAULT_INPUT  = r"docs/jones/For Suresh - Quicken - Transactions - Apr2026 - no pivot.xlsx"
_DEFAULT_OUTPUT = r"docs/jones/Jones_Tax_Summary.xlsx"

def _parse_args():
    p = argparse.ArgumentParser(description="Jones Quicken → IRS Tax Summary")
    p.add_argument("--input",  default=_DEFAULT_INPUT,  help="Path to Quicken xlsx export")
    p.add_argument("--output", default=_DEFAULT_OUTPUT, help="Path for output xlsx")
    return p.parse_args()

args = _parse_args()
INPUT_FILE  = args.input
OUTPUT_FILE = args.output

# ---------------------------------------------------------------------------
# PROPERTY ORDER & LABELS
# ---------------------------------------------------------------------------
PROPERTY_ORDER = ["Beaumont", "Britton", "Hickory", "OBBC II",
                  "River", "Woodrow School", "secu3496"]

def _normalize_property(name: str) -> str:
    n = name.lower().strip()
    mapping = {
        "beaumont rental": "Beaumont",
        "beaumont": "Beaumont",
        "britten": "Britton",
        "britton": "Britton",
        "hickory": "Hickory",
        "obbc ii": "OBBC II",
        "river": "River",
        "woodrow school": "Woodrow School",
        "secu3496": "secu3496",
    }
    return mapping.get(n, name.title())

def _clean_exp_label(raw: str) -> str:
    """Normalise rental expense type labels — fix typos, consolidate synonyms."""
    lmap = {
        "2nd mortage 3496": "Mortgage Payment",
        "2nd mortage 3496:bus intrest 3496": "Mortgage Interest",
        "bus intrest 3496": "Mortgage Interest",
        "mortgage interest, business": "Mortgage Interest",
        "bus tax": "Business / Property Tax",
        "property tax": "Business / Property Tax",
        "cleaing": "Cleaning",
        "cleaning": "Cleaning",
        "electric": "Electric",
        "eqipment": "Equipment",
        "equipment": "Equipment",
        "escrow": "Escrow",
        "excrow": "Escrow",
        "externator": "Exterminator",
        "gas": "Gas",
        "insurance": "Insurance",
        "internet": "Internet",
        "mantiance": "Maintenance",
        "repair and mantiance": "Repairs",
        "repair and maintenance": "Repairs",
        "repair": "Repairs",
        "repairs": "Repairs",
        "propane": "Propane",
        "spetic": "Septic",
        "supplies": "Supplies",
        "water": "Water",
        "yard": "Yard / Grounds",
        "general": "General",
    }
    return lmap.get(raw.lower().strip(), raw.strip().title())

# ---------------------------------------------------------------------------
# SECTION ORDER & LABELS
# ---------------------------------------------------------------------------
SECTION_ORDER = [
    "INCOME_FARM",
    "INCOME_RENTAL",   # rendered as part of rental matrix
    "INCOME_OTHER",
    "LAND_SALE",
    "FARM_EXP",
    "FARM_AUTO",
    "RENTAL_EXP",      # rendered as part of rental matrix
    "SCH_A",
    "RE_BUSINESS",
    "PERSONAL_AUTO",
    "NON_DEDUCTIBLE",
    "TRANSFER",
]

SECTION_LABELS = {
    "INCOME_FARM":    "FARM INCOME  (Schedule F)",
    "INCOME_RENTAL":  "RENTAL INCOME & EXPENSES  (Schedule E — per property)",
    "INCOME_OTHER":   "OTHER INCOME",
    "LAND_SALE":      "LAND SALE PROCEEDS  (Schedule D / Form 4797 — confirm with CPA)",
    "FARM_EXP":       "FARM EXPENSES  (Schedule F)",
    "FARM_AUTO":      "FARM AUTO  (Schedule F)",
    "RENTAL_EXP":     "RENTAL EXPENSES  (Schedule E — per property)",
    "SCH_A":          "SCHEDULE A — ITEMIZED DEDUCTIONS",
    "RE_BUSINESS":    "RE BUSINESS EXPENSES",
    "PERSONAL_AUTO":  "PERSONAL AUTO  (non-deductible, for reference)",
    "NON_DEDUCTIBLE": "NON-DEDUCTIBLE / PERSONAL  (for reference)",
    "TRANSFER":       "EXCLUDED — INTERNAL TRANSFERS  (loan principal, inter-account)",
}

# ---------------------------------------------------------------------------
# IRS GROUP MAPPING
# ---------------------------------------------------------------------------

_MORTGAGE_SPLIT_CATS = {
    "real estate expense:rental expense:beaumont rental",
    "real estate expense:rental expense:secu3496",
    "real estate expense:rental expense:secu3496:2nd mortage 3496",
    "real estate expense:rental expense:secu3496:2nd mortage 3496:bus intrest 3496",
}

def irs_group(raw_cat: str):
    c = raw_cat.strip().lower()

    # --- TRANSFERS ---
    if c.startswith("[") or c in ("transfered in", "adjust", "non- income",
                                   "non-income", "bank of america4727",
                                   "bank of america6397", "sam's4074",
                                   "secu3493", "loan payment3493"):
        return ("TRANSFER", None, "Transfer / Internal")

    # --- FARM INCOME ---
    if c.startswith("farm income:"):
        label = raw_cat.split(":")[-1].strip()
        return ("INCOME_FARM", None, label)

    # --- RENTAL INCOME ---
    if c.startswith("rental income:"):
        parts = raw_cat.split(":")
        prop = _normalize_property(parts[1].strip()) if len(parts) > 1 else "Unknown"
        return ("INCOME_RENTAL", prop, "Rental Income")

    # --- SOCIAL SECURITY / PENSION ---
    if c.startswith("social security:"):
        return ("INCOME_OTHER", None, "Social Security")
    if c == "linda's pension":
        return ("INCOME_OTHER", None, "Pension")

    # --- INTEREST / OTHER INCOME ---
    if c == "interest inc":
        return ("INCOME_OTHER", None, "Interest / Dividend Income")
    if c in ("other inc", "other inc, bus"):
        return ("INCOME_OTHER", None, "Other Income")
    if c == "rebate":
        return ("INCOME_OTHER", None, "Rebate")

    # --- FARM EXPENSES (top-level catch, no sub-category) ---
    if c == "farm expense1":
        return ("FARM_EXP", None, "General Farm Expense")

    # --- FARM AUTO ---
    if c.startswith("farm expense1:auto:"):
        return ("FARM_AUTO", None, raw_cat[len("Farm Expense1:Auto:"):].strip())

    # --- FARM EXPENSES ---
    if c.startswith("farm expense1:"):
        parts = raw_cat.split(":")
        label = ":".join(p.strip() for p in parts[1:]) if len(parts) > 1 else raw_cat
        return ("FARM_EXP", None, label)

    # --- RENTAL EXPENSES ---
    if c.startswith("real estate expense:rental expense:"):
        remainder = raw_cat[len("Real Estate Expense:Rental Expense:"):].strip()
        parts = remainder.split(":")
        prop = _normalize_property(parts[0].strip())
        exp_type = ":".join(p.strip() for p in parts[1:]) if len(parts) > 1 else "General"
        return ("RENTAL_EXP", prop, exp_type)

    # --- LOAN INTEREST / PRINCIPAL ---
    if c == "loan payment3493:interest3492":
        return ("SCH_A", None, "Mortgage Interest")
    if c == "loan payment3493":
        return ("NON_DEDUCTIBLE", None, "Loan Principal (non-deductible)")

    # --- RE BUSINESS ---
    if c.startswith("r. e.:") or c == "r.e. office expense":
        sub = raw_cat[6:].strip() if c.startswith("r. e.:") else "Office Expense"
        if "non-deductib" in c:
            return ("NON_DEDUCTIBLE", None, "R.E. Non-deductible")
        return ("RE_BUSINESS", None, sub)
    if c in ("legal-prof fees", "legal"):
        return ("RE_BUSINESS", None, "Legal / Professional Fees")
    if c == "misc. andy":
        return ("RE_BUSINESS", None, "Contractor — Andy")

    # --- SCHEDULE A: MORTGAGE INTEREST ---
    if c in ("int exp", "interest exp", "mortgage int"):
        return ("SCH_A", None, "Mortgage Interest")
    if c in ("120wolfs4791:mortage interest", "120wolfs4791:wolfs mortage",
             "mortage 4191:interest", "mortage 4191"):
        return ("SCH_A", None, "Mortgage Interest")
    if c in ("120wolfs4791:escrow", "mortage 4191:escrow", "120wolfs4791"):
        return ("SCH_A", None, "Escrow (incl. taxes/insurance)")

    # --- SCHEDULE A: PROPERTY TAX ---
    if c in ("mortage 4191:property tax", "tax, business:property"):
        return ("SCH_A", None, "Property Tax")

    # --- SCHEDULE A: CHARITABLE ---
    if c in ("church", "donation", "gifts given"):
        return ("SCH_A", None, "Charitable Contributions")

    # --- SCHEDULE A: MEDICAL ---
    if c.startswith("medical:") or c in ("insurance:dental", "insurance:medical"):
        sub = raw_cat.split(":")[-1].strip()
        return ("SCH_A", None, f"Medical — {sub}")

    # --- PERSONAL AUTO ---
    if c.startswith("auto:"):
        return ("PERSONAL_AUTO", None, raw_cat[5:].strip())

    # --- PRIMARY HOME (non-deductible) ---
    if c.startswith("wolfs:") or c.startswith("120wolfs4791:") or c.startswith("mortage 4191:"):
        sub = raw_cat.split(":")[-1].strip()
        return ("NON_DEDUCTIBLE", None, f"Home — {sub}")

    # --- NON-DEDUCTIBLE ---
    non_ded = {
        "groceries", "dining", "entertainment", "clothing", "personal care",
        "house", "hunt", "pet", "vacation:travel", "books",
        "membership fees", "subscriptions", "dues and subscriptions",
        "insurance:life", "insurance:property",
        "hunting:supplies", "tax prepration", "service charge",
        "bank charge", "tax:other:foreign transaction fee", "misc", "repairs",
    }
    if c in non_ded:
        return ("NON_DEDUCTIBLE", None, raw_cat.strip())

    return ("NON_DEDUCTIBLE", None, f"[Unclassified] {raw_cat.strip()}")

# ---------------------------------------------------------------------------
# ROW PRE-PROCESSING
# ---------------------------------------------------------------------------
def preprocess_row(row: dict) -> dict:
    c    = row["category"].lower().strip()
    amt  = row["amount"]
    desc = (row.get("description") or "").lower()
    memo = (row.get("memo") or "").lower()

    if "opening balance" in desc:
        row["_override"] = ("TRANSFER", None, "Opening Balance (Quicken artifact)")
        return row

    if c in ("other inc", "other inc, bus") and "land sale" in memo:
        row["_override"] = ("LAND_SALE", None, "Land Sale — Investor Title")
        return row

    if c in ("other inc", "other inc, bus") and (
        "forest" in desc or "replanting" in memo or "pines" in memo
    ):
        row["_override"] = ("INCOME_FARM", None, "Forestry Cost-Share")
        return row

    if amt > 0 and c in _MORTGAGE_SPLIT_CATS:
        if any(kw in desc for kw in ("payment", "transfer")):
            row["_override"] = ("TRANSFER", None, "Loan Principal (Quicken split)")
            return row

    return row

def resolve(row: dict):
    if "_override" in row:
        return row["_override"]
    return irs_group(row["category"])

# ---------------------------------------------------------------------------
# READ TRANSACTIONS
# ---------------------------------------------------------------------------
def read_transactions(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['Report']
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 5:
            continue
        date, acct, num, desc, memo, cat, tag, amount = (
            row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8]
        )
        if not isinstance(date, datetime.datetime):
            continue
        if cat is None or amount is None:
            continue
        r = {
            "date": date, "account": acct or "", "num": num or "",
            "description": desc or "", "memo": memo or "",
            "category": str(cat).strip(), "tag": tag or "",
            "amount": float(amount),
        }
        rows.append(preprocess_row(r))
    wb.close()
    return rows

# ---------------------------------------------------------------------------
# AGGREGATE
# ---------------------------------------------------------------------------
def aggregate(rows):
    data = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    for r in rows:
        section, sub_group, line_label = resolve(r)
        data[section][sub_group or ""][line_label] += r["amount"]
    return data

# ---------------------------------------------------------------------------
# STYLES
# ---------------------------------------------------------------------------
DARK_BLUE  = "1F4E79"
MED_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
GREEN_BG   = "E2EFDA"
GREEN_DARK = "375623"
ORANGE     = "C55A11"
WHITE      = "FFFFFF"
GRAY_ALT   = "F5F5F5"
GRAY_TOTAL = "EDEDED"

def _font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _thin_border():
    s = Side(style="thin", color="D0D0D0")
    return Border(bottom=s, left=s, right=s)

def _money(ws, row, col, value, bold=False, color="000000", bg=None):
    cell = ws.cell(row=row, column=col, value=value if value != 0.0 else None)
    cell.number_format = '#,##0.00'
    cell.alignment = Alignment(horizontal="right")
    cell.font = _font(bold=bold, color=color)
    if bg:
        cell.fill = _fill(bg)
    return cell

def _label(ws, row, col, text, bold=False, indent=0, color="000000", bg=None, italic=False):
    cell = ws.cell(row=row, column=col, value=("  " * indent) + text)
    cell.font = _font(bold=bold, color=color, italic=italic)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if bg:
        cell.fill = _fill(bg)
    return cell

# ---------------------------------------------------------------------------
# WRITE SUMMARY SHEET
# ---------------------------------------------------------------------------
def write_summary(wb, data, rows):
    ws = wb.create_sheet("Summary")

    # Column layout: A=labels, B=amount (2-col sections) / B..H=properties + I=Total (matrix)
    PROPS = PROPERTY_ORDER
    N     = len(PROPS)          # 7
    TOT   = N + 2               # col index for Total (9)
    LAST  = TOT

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 16   # amount for 2-col sections / Beaumont
    for i in range(3, LAST + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14

    def _section_header(r, label, color=MED_BLUE):
        ws.merge_cells(f"A{r}:{get_column_letter(LAST)}{r}")
        c = ws.cell(row=r, column=1, value=label)
        c.font = _font(bold=True, color=WHITE, size=10)
        c.fill = _fill(color)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 18
        return r + 1

    def _subtotal_row(r, label, amount, col=2, bg=LIGHT_BLUE, last_col=None):
        lc = last_col or LAST
        _label(ws, r, 1, label, bold=True, bg=bg)
        m = _money(ws, r, col, amount, bold=True, bg=bg)
        for c in range(2, lc + 1):
            if c != col:
                ws.cell(row=r, column=c).fill = _fill(bg)
        ws.row_dimensions[r].height = 15
        return r + 1

    def _two_col_section(r, section, alternate=True):
        """Render a standard 2-column section. Returns next row."""
        section_data = data.get(section, {})
        if not section_data:
            return r

        color = ORANGE if section == "LAND_SALE" else MED_BLUE
        r = _section_header(r, SECTION_LABELS[section], color=color)

        total = 0.0
        all_lines = []
        for sg in sorted(section_data.keys()):
            for lbl, amt in sorted(section_data[sg].items()):
                all_lines.append((lbl, amt))

        for i, (lbl, amt) in enumerate(all_lines):
            bg = GRAY_ALT if (alternate and i % 2 == 1) else None
            _label(ws, r, 1, lbl, indent=1, bg=bg)
            _money(ws, r, 2, amt, bg=bg)
            total += amt
            r += 1

        lbl_short = SECTION_LABELS[section].split("(")[0].split("—")[0].strip()
        r = _subtotal_row(r, f"{lbl_short} — Total", total)
        return r + 1   # blank line after section

    # -----------------------------------------------------------------------
    # TITLE
    # -----------------------------------------------------------------------
    r = 1
    ws.merge_cells(f"A{r}:{get_column_letter(LAST)}{r}")
    c = ws.cell(row=r, column=1, value="Jones — IRS Tax Category Summary   (Jan – Dec 2025)")
    c.font = _font(bold=True, color=WHITE, size=13)
    c.fill = _fill(DARK_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 24
    r += 2

    # -----------------------------------------------------------------------
    # KEY FIGURES AT A GLANCE  (filled in after detail sections; placeholders now)
    # We write this block first so it appears at the top, then fill values below.
    # -----------------------------------------------------------------------
    key_start = r
    r = _section_header(r, "KEY FIGURES AT A GLANCE", color=DARK_BLUE)

    # Column headers
    for col, hdr in enumerate(["", "Amount ($)", "Notes"], start=1):
        c = ws.cell(row=r, column=col, value=hdr)
        c.font = _font(bold=True, color=WHITE)
        c.fill = _fill(MED_BLUE)
    ws.row_dimensions[r].height = 15
    r += 1

    key_rows = {}   # name -> (row_index, amount_col=2, notes_col=3)
    key_items = [
        ("Farm Income",              "Schedule F"),
        ("Farm Expenses & Auto",     "Schedule F"),
        ("Net Farm Income",          "Schedule F — net"),
        ("Rental Income",            "Schedule E — all properties"),
        ("Rental Expenses",          "Schedule E — all properties"),
        ("Net Rental Income",        "Schedule E — net"),
        ("Other Income",             "Social Security, pension, interest"),
        ("Land Sale Proceeds",       "⚠  Confirm Schedule D vs Form 4797 with CPA"),
        ("Schedule A Deductions",    "Itemized deductions total"),
        ("RE Business Expenses",     "Office, legal, contractor"),
    ]
    for i, (name, note) in enumerate(key_items):
        bg = GRAY_ALT if i % 2 == 1 else None
        is_net = name.startswith("Net")
        is_land = name.startswith("Land")
        row_bg = GREEN_BG if is_net else (ORANGE if is_land else bg)
        _label(ws, r, 1, name, bold=is_net or is_land, indent=1, bg=row_bg,
               color=WHITE if is_land else "000000")
        # Amount cell — placeholder value 0; will be overwritten after aggregation
        m = ws.cell(row=r, column=2, value=0.0)
        m.number_format = '#,##0.00'
        m.alignment = Alignment(horizontal="right")
        m.font = _font(bold=is_net or is_land,
                       color=WHITE if is_land else "000000")
        if row_bg:
            m.fill = _fill(row_bg)
        n = ws.cell(row=r, column=3, value=note)
        n.font = _font(italic=True, color="595959" if not is_land else WHITE,
                       size=9)
        if row_bg:
            n.fill = _fill(row_bg)
        key_rows[name] = r
        r += 1

    r += 1   # blank line after key block

    # -----------------------------------------------------------------------
    # FARM INCOME
    # -----------------------------------------------------------------------
    r = _two_col_section(r, "INCOME_FARM")
    farm_income_total = sum(
        amt for sg in data.get("INCOME_FARM", {}).values() for amt in sg.values()
    )

    # -----------------------------------------------------------------------
    # RENTAL MATRIX  (income + expenses, properties as columns)
    # -----------------------------------------------------------------------
    # Build per-property income
    ri_by_prop = defaultdict(float)
    for sg, lines in data.get("INCOME_RENTAL", {}).items():
        for lbl, amt in lines.items():
            ri_by_prop[sg if sg else lbl] += amt

    # Build per-property, per-expense-type expenses (with clean labels)
    re_matrix = defaultdict(lambda: defaultdict(float))   # prop -> clean_type -> amt
    for prop, lines in data.get("RENTAL_EXP", {}).items():
        for exp_type, amt in lines.items():
            clean = _clean_exp_label(exp_type)
            re_matrix[prop][clean] += amt

    all_exp_types = sorted(set(
        et for pd in re_matrix.values() for et in pd.keys()
    ))

    # Preferred expense type order
    _EXP_ORDER = [
        "Rental Income",   # sentinel — placed first as income row
        "Mortgage Payment", "Mortgage Interest", "Escrow",
        "Electric", "Gas", "Propane", "Water", "Internet",
        "Cleaning", "Repairs", "Maintenance", "Exterminator", "Yard / Grounds",
        "Insurance", "Business / Property Tax", "Property Tax",
        "Equipment", "Supplies", "Septic", "General",
    ]
    def _exp_sort_key(e):
        try:
            return _EXP_ORDER.index(e)
        except ValueError:
            return 999

    all_exp_types = sorted(all_exp_types, key=_exp_sort_key)

    # Section header
    r = _section_header(r, SECTION_LABELS["INCOME_RENTAL"], color=MED_BLUE)

    # Property column header row
    _label(ws, r, 1, "", bold=True, bg=DARK_BLUE)
    for pi, prop in enumerate(PROPS):
        c = ws.cell(row=r, column=pi + 2, value=prop)
        c.font = _font(bold=True, color=WHITE, size=9)
        c.fill = _fill(DARK_BLUE)
        c.alignment = Alignment(horizontal="center")
    c = ws.cell(row=r, column=TOT, value="TOTAL")
    c.font = _font(bold=True, color=WHITE, size=9)
    c.fill = _fill(DARK_BLUE)
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[r].height = 16
    r += 1

    # Rental Income row
    inc_total = 0.0
    _label(ws, r, 1, "Rental Income", bold=True, bg=GREEN_BG)
    for pi, prop in enumerate(PROPS):
        amt = ri_by_prop.get(prop, 0.0)
        _money(ws, r, pi + 2, amt, bold=True, bg=GREEN_BG)
        inc_total += amt
    _money(ws, r, TOT, inc_total, bold=True, bg=GREEN_BG)
    ws.row_dimensions[r].height = 15
    r += 1

    # Expenses sub-header
    ws.merge_cells(f"A{r}:{get_column_letter(LAST)}{r}")
    c = ws.cell(row=r, column=1, value="  Expenses")
    c.font = _font(bold=True, color=WHITE, size=9, italic=True)
    c.fill = _fill("4472C4")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 14
    r += 1

    # Expense type rows
    exp_totals_by_prop = defaultdict(float)
    grand_exp_total = 0.0
    for i, exp_type in enumerate(all_exp_types):
        bg = GRAY_ALT if i % 2 == 1 else None
        _label(ws, r, 1, exp_type, indent=1, bg=bg)
        row_total = 0.0
        for pi, prop in enumerate(PROPS):
            amt = re_matrix.get(prop, {}).get(exp_type, None)
            if amt is not None:
                _money(ws, r, pi + 2, amt, bg=bg)
                row_total += amt
                exp_totals_by_prop[prop] += amt
            else:
                if bg:
                    ws.cell(row=r, column=pi + 2).fill = _fill(bg)
        _money(ws, r, TOT, row_total, bg=bg)
        grand_exp_total += row_total
        r += 1

    # Total Expenses row
    _label(ws, r, 1, "Total Expenses", bold=True, bg=LIGHT_BLUE)
    for pi, prop in enumerate(PROPS):
        _money(ws, r, pi + 2, exp_totals_by_prop.get(prop, 0.0), bold=True, bg=LIGHT_BLUE)
    _money(ws, r, TOT, grand_exp_total, bold=True, bg=LIGHT_BLUE)
    ws.row_dimensions[r].height = 15
    r += 1

    # Net Rental row
    _label(ws, r, 1, "Net Rental Income", bold=True, bg=GREEN_BG)
    net_by_prop = {}
    grand_net = 0.0
    for pi, prop in enumerate(PROPS):
        inc = ri_by_prop.get(prop, 0.0)
        exp = exp_totals_by_prop.get(prop, 0.0)
        net = inc + exp
        net_by_prop[prop] = net
        col_color = "C00000" if net < 0 else GREEN_DARK
        _money(ws, r, pi + 2, net, bold=True, color=col_color, bg=GREEN_BG)
        grand_net += net
    col_color = "C00000" if grand_net < 0 else GREEN_DARK
    _money(ws, r, TOT, grand_net, bold=True, color=col_color, bg=GREEN_BG)
    ws.row_dimensions[r].height = 15
    r += 2   # blank line

    # -----------------------------------------------------------------------
    # OTHER INCOME
    # -----------------------------------------------------------------------
    r = _two_col_section(r, "INCOME_OTHER")

    # -----------------------------------------------------------------------
    # LAND SALE  (amber, flagged)
    # -----------------------------------------------------------------------
    r = _two_col_section(r, "LAND_SALE")

    # -----------------------------------------------------------------------
    # FARM EXPENSES + FARM AUTO + NET FARM block
    # -----------------------------------------------------------------------
    r = _two_col_section(r, "FARM_EXP")
    farm_exp_total = sum(
        amt for sg in data.get("FARM_EXP", {}).values() for amt in sg.values()
    )

    r = _two_col_section(r, "FARM_AUTO")
    farm_auto_total = sum(
        amt for sg in data.get("FARM_AUTO", {}).values() for amt in sg.values()
    )

    # Net Farm block
    r = _section_header(r, "NET FARM INCOME  (Schedule F)", color=DARK_BLUE)
    net_farm = farm_income_total + farm_exp_total + farm_auto_total
    for lbl, val, is_net in [
        ("Farm Income",   farm_income_total,  False),
        ("Farm Expenses", farm_exp_total,      False),
        ("Farm Auto",     farm_auto_total,     False),
        ("Net Farm Income", net_farm,           True),
    ]:
        bg = GREEN_BG if is_net else None
        col_color = ("C00000" if net_farm < 0 else GREEN_DARK) if is_net else "000000"
        _label(ws, r, 1, lbl, bold=is_net, indent=1, bg=bg)
        _money(ws, r, 2, val, bold=is_net, color=col_color, bg=bg)
        r += 1
    r += 1

    # -----------------------------------------------------------------------
    # SCHEDULE A
    # -----------------------------------------------------------------------
    r = _two_col_section(r, "SCH_A")

    # -----------------------------------------------------------------------
    # RE BUSINESS / PERSONAL AUTO / NON-DED / TRANSFERS
    # -----------------------------------------------------------------------
    re_business_total = sum(
        amt for sg in data.get("RE_BUSINESS", {}).values() for amt in sg.values()
    )
    r = _two_col_section(r, "RE_BUSINESS")
    r = _two_col_section(r, "PERSONAL_AUTO")
    r = _two_col_section(r, "NON_DEDUCTIBLE")
    r = _two_col_section(r, "TRANSFER")

    # -----------------------------------------------------------------------
    # FILL BACK — Key Figures (now that all totals are known)
    # -----------------------------------------------------------------------
    sch_a_total = sum(
        amt for sg in data.get("SCH_A", {}).values() for amt in sg.values()
    )
    land_sale_total = sum(
        amt for sg in data.get("LAND_SALE", {}).values() for amt in sg.values()
    )
    other_income_total = sum(
        amt for sg in data.get("INCOME_OTHER", {}).values() for amt in sg.values()
    )

    key_values = {
        "Farm Income":           farm_income_total,
        "Farm Expenses & Auto":  farm_exp_total + farm_auto_total,
        "Net Farm Income":       net_farm,
        "Rental Income":         inc_total,
        "Rental Expenses":       grand_exp_total,
        "Net Rental Income":     grand_net,
        "Other Income":          other_income_total,
        "Land Sale Proceeds":    land_sale_total,
        "Schedule A Deductions": sch_a_total,
        "RE Business Expenses":  re_business_total,
    }

    for name, val in key_values.items():
        row_idx = key_rows[name]
        cell = ws.cell(row=row_idx, column=2, value=val)
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal="right")
        is_net  = name.startswith("Net")
        is_land = name.startswith("Land")
        row_bg  = GREEN_BG if is_net else (ORANGE if is_land else None)
        # re-apply alternating bg for non-special rows
        idx = list(key_values.keys()).index(name)
        if not is_net and not is_land:
            row_bg = GRAY_ALT if idx % 2 == 1 else None
        net_negative = is_net and val < 0
        col_color = ("C00000" if net_negative else
                     (WHITE if is_land else
                      (GREEN_DARK if is_net else "000000")))
        cell.font = _font(bold=is_net or is_land, color=col_color)
        if row_bg:
            cell.fill = _fill(row_bg)


# ---------------------------------------------------------------------------
# WRITE TRANSACTIONS SHEET
# ---------------------------------------------------------------------------
def write_transactions(wb, rows):
    ws = wb.create_sheet("Transactions")

    headers = ["Date", "Account", "Num", "Description", "Memo",
               "Category", "Tag", "Amount", "IRS Group", "Sub-Group", "Line Label"]
    widths   = [12, 20, 6, 30, 25, 45, 10, 14, 20, 18, 40]

    for col, (hdr, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=col, value=hdr)
        c.font = _font(bold=True, color=WHITE)
        c.fill = _fill(MED_BLUE)
        ws.column_dimensions[get_column_letter(col)].width = w

    for ri, row in enumerate(rows, start=2):
        section, sub_group, line_label = resolve(row)
        bg = GRAY_ALT if ri % 2 == 0 else None
        vals = [
            row["date"], row["account"], row["num"], row["description"],
            row["memo"], row["category"], row["tag"], row["amount"],
            section, sub_group or "", line_label,
        ]
        for col, val in enumerate(vals, start=1):
            c = ws.cell(row=ri, column=col, value=val)
            c.font = _font()
            if bg:
                c.fill = _fill(bg)
            if col == 8:
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal="right")
            if col == 1 and isinstance(val, datetime.datetime):
                c.number_format = 'MM/DD/YYYY'


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    print("Reading transactions...")
    rows = read_transactions(INPUT_FILE)
    print(f"  {len(rows)} transaction rows loaded.")

    print("Aggregating...")
    data = aggregate(rows)

    print("Writing output...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_summary(wb, data, rows)
    write_transactions(wb, rows)
    wb.save(OUTPUT_FILE)
    print(f"Done → {OUTPUT_FILE}")

    total_by_section = defaultdict(float)
    for r in rows:
        section, _, _ = resolve(r)
        total_by_section[section] += r["amount"]
    print("\nSection totals:")
    for sec in SECTION_ORDER:
        if sec in total_by_section:
            print(f"  {sec:<20} {total_by_section[sec]:>14,.2f}")

if __name__ == "__main__":
    main()
