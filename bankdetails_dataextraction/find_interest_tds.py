#!/usr/bin/env python3
"""
find_interest_tds.py — Find Interest Income and Tax Deducted transactions
from bank transaction Excel/CSV files.

CLI usage:
    python find_interest_tds.py transactions.csv
    python find_interest_tds.py transactions.xlsx --locale us --threshold 0.6
    python find_interest_tds.py transactions.xlsx --output results.xlsx

Importable:
    from find_interest_tds import find_transactions
    result = find_transactions("transactions.xlsx", locale="india")
    # result = { "interest_income": DataFrame, "tax_deducted": DataFrame }
"""

import os
os.environ.setdefault("HF_HUB_DISABLE_SYMLINKS_WARNING", "1")

import argparse
import re
from pathlib import Path

import numpy as np
import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

CONFIG_DIR = Path(__file__).parent / "interest_tds_configs"
MODEL_NAME = "all-MiniLM-L6-v2"

# Header colour: steel blue (matches reconciliation sheet style)
HEADER_FILL = PatternFill(fill_type="solid", fgColor="2F5496")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TOTAL_FONT = Font(bold=True)


# ---------------------------------------------------------------------------
# Config + file loading
# ---------------------------------------------------------------------------

def _load_config(locale: str) -> dict:
    config_path = CONFIG_DIR / f"{locale}.yaml"
    if not config_path.exists():
        available = [p.stem for p in CONFIG_DIR.glob("*.yaml")]
        raise FileNotFoundError(
            f"No config for locale '{locale}'. Available: {available}"
        )
    with open(config_path, encoding="utf-8") as f:
        return yaml.safe_load(f)


def _load_file(filepath: str) -> pd.DataFrame:
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {filepath}")
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        return pd.read_excel(filepath, dtype=str)
    if suffix == ".csv":
        return pd.read_csv(filepath, dtype=str)
    raise ValueError(f"Unsupported file type: {path.suffix}. Use .xlsx, .xls, or .csv")


# ---------------------------------------------------------------------------
# Column detection
# ---------------------------------------------------------------------------

def _detect_column(df_cols: list, hints: list):
    lower_cols = [c.lower().strip() for c in df_cols]
    # Exact match first
    for hint in hints:
        hl = hint.lower()
        if hl in lower_cols:
            return df_cols[lower_cols.index(hl)]
    # Substring match
    for hint in hints:
        hl = hint.lower()
        for i, col in enumerate(lower_cols):
            if hl in col or col in hl:
                return df_cols[i]
    return None


def _detect_columns(df: pd.DataFrame, config: dict):
    hints = config.get("column_hints", {})
    cols = list(df.columns)

    date_col   = _detect_column(cols, hints.get("date", []))
    desc_col   = _detect_column(cols, hints.get("description", []))
    amount_col = _detect_column(cols, hints.get("amount", []))

    missing = [name for name, val in
               [("date", date_col), ("description", desc_col), ("amount", amount_col)]
               if val is None]
    if missing:
        raise ValueError(
            f"Could not detect columns for: {missing}.\n"
            f"File columns: {cols}\n"
            f"Add matching hints to the YAML config."
        )
    return date_col, desc_col, amount_col


# ---------------------------------------------------------------------------
# Pass 1 — keyword regex
# ---------------------------------------------------------------------------

def _keyword_pass(df: pd.DataFrame, desc_col: str, categories: dict) -> pd.Series:
    labels = pd.Series([None] * len(df), index=df.index, dtype=object)
    for category, cfg in categories.items():
        keywords = cfg.get("keywords", [])
        if not keywords:
            continue
        pattern = "|".join(rf"\b{re.escape(kw)}\b" for kw in keywords)
        matched = df[desc_col].fillna("").str.contains(pattern, case=False, regex=True)
        # Only label rows not yet claimed
        labels[(matched) & (labels.isna())] = category
    return labels


# ---------------------------------------------------------------------------
# Pass 2 — cosine similarity
# ---------------------------------------------------------------------------

def _cosine_pass(
    df: pd.DataFrame,
    desc_col: str,
    categories: dict,
    threshold: float,
    review_threshold: float,
    model,
):
    """
    Returns three Series (all indexed like df):
      labels        — category name for rows >= threshold
      scores        — cosine score for matched rows
      review_labels — category of nearest anchor for rows in [review_threshold, threshold)
      review_scores — cosine score for near-miss rows
    """
    from sklearn.metrics.pairwise import cosine_similarity as cos_sim

    all_anchors, anchor_cats = [], []
    for category, cfg in categories.items():
        for anchor in cfg.get("anchors", []):
            all_anchors.append(anchor)
            anchor_cats.append(category)

    empty = pd.Series([None] * len(df), index=df.index, dtype=object)
    if not all_anchors or df.empty:
        return empty.copy(), empty.copy(), empty.copy(), empty.copy()

    descriptions = df[desc_col].fillna("").tolist()
    desc_emb   = model.encode(descriptions, show_progress_bar=False)
    anchor_emb = model.encode(all_anchors,  show_progress_bar=False)
    sim_matrix = cos_sim(desc_emb, anchor_emb)  # (n_rows, n_anchors)

    labels        = empty.copy()
    scores        = empty.copy()
    review_labels = empty.copy()
    review_scores = empty.copy()

    for df_idx, row_scores in zip(df.index, sim_matrix):
        best      = int(np.argmax(row_scores))
        best_score = float(row_scores[best])
        if best_score >= threshold:
            labels[df_idx] = anchor_cats[best]
            scores[df_idx] = round(best_score, 2)
        elif best_score >= review_threshold:
            review_labels[df_idx] = anchor_cats[best]
            review_scores[df_idx] = round(best_score, 2)

    return labels, scores, review_labels, review_scores


# ---------------------------------------------------------------------------
# Output writing
# ---------------------------------------------------------------------------

REVIEW_FILL = PatternFill(fill_type="solid", fgColor="BF8F00")  # amber — signals "needs review"


def _write_sheet(ws, df: pd.DataFrame, col_widths: dict, header_fill, with_total: bool):
    headers = list(df.columns)

    for ci, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=header)
        cell.font = HEADER_FONT
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, value in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=value)

    if with_total and "Debit/Credit" in headers:
        total_row = len(df) + 2
        amt_col_idx = headers.index("Debit/Credit") + 1
        ws.cell(row=total_row, column=1, value="Total").font = TOTAL_FONT
        total_cell = ws.cell(row=total_row, column=amt_col_idx)
        total_cell.value = round(float(df["Debit/Credit"].sum()), 2)
        total_cell.font = TOTAL_FONT

    for ci, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(header, 15)


def _write_output(result: dict, output_path: str, config: dict):
    wb = Workbook()
    wb.remove(wb.active)

    category_display = {
        cat: cfg.get("display_name", cat.replace("_", " ").title())
        for cat, cfg in config.get("categories", {}).items()
    }

    main_col_widths   = {"Date": 15, "Description": 48, "Debit/Credit": 18, "Match": 22}
    review_col_widths = {"Date": 15, "Description": 48, "Debit/Credit": 18,
                         "Nearest Category": 20, "Score": 10}

    for category, df in result.items():
        if category == "_review":
            if df.empty:
                continue
            ws = wb.create_sheet(title="Near Miss — Review")
            _write_sheet(ws, df, review_col_widths, REVIEW_FILL, with_total=False)
        else:
            sheet_title = category_display.get(category, category)[:31]
            ws = wb.create_sheet(title=sheet_title)
            _write_sheet(ws, df, main_col_widths, HEADER_FILL, with_total=True)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def find_transactions(
    filepath: str,
    locale: str = "india",
    threshold: float = None,
    output: str = None,
) -> dict:
    """
    Find interest income and tax deducted transactions in a bank statement file.

    Args:
        filepath  : Path to .xlsx, .xls, or .csv file
        locale    : YAML config to use — 'india' or 'us' (extensible)
        threshold : Cosine similarity cutoff (0–1). Overrides YAML value if provided.
        output    : Output Excel path. Defaults to <input>_interest_tds.xlsx

    Returns:
        dict[category_name -> DataFrame] with columns Date | Description | Debit/Credit | Match
        Plus key '_review' -> DataFrame of near-miss rows
    """
    config     = _load_config(locale)
    df         = _load_file(filepath)
    categories = config.get("categories", {})

    # Thresholds: CLI arg overrides YAML, YAML overrides hardcoded default
    main_threshold   = threshold if threshold is not None else config.get("threshold", 0.55)
    review_threshold = config.get("review_threshold", 0.35)

    date_col, desc_col, amount_col = _detect_columns(df, config)

    # Pass 1 — keywords
    kw_labels = _keyword_pass(df, desc_col, categories)

    # Pass 2 — cosine on unmatched rows only
    unmatched_mask = kw_labels.isna()
    cos_labels        = pd.Series([None] * len(df), index=df.index, dtype=object)
    cos_scores        = pd.Series([None] * len(df), index=df.index, dtype=object)
    review_labels_all = pd.Series([None] * len(df), index=df.index, dtype=object)
    review_scores_all = pd.Series([None] * len(df), index=df.index, dtype=object)

    if unmatched_mask.any():
        print(f"Loading model '{MODEL_NAME}' for semantic matching...")
        from sentence_transformers import SentenceTransformer
        model = SentenceTransformer(MODEL_NAME)
        sub_labels, sub_scores, sub_rev_labels, sub_rev_scores = _cosine_pass(
            df[unmatched_mask], desc_col, categories, main_threshold, review_threshold, model
        )
        cos_labels[unmatched_mask]        = sub_labels.values
        cos_scores[unmatched_mask]        = sub_scores.values
        review_labels_all[unmatched_mask] = sub_rev_labels.values
        review_scores_all[unmatched_mask] = sub_rev_scores.values

    # Combine labels + build Match audit column
    final_labels = kw_labels.combine_first(cos_labels)
    match_col = pd.Series([None] * len(df), index=df.index, dtype=object)
    match_col[kw_labels.notna()] = "keyword"
    cosine_hits = cos_labels.notna()
    match_col[cosine_hits] = cos_scores[cosine_hits].apply(lambda s: f"cosine ({s})")

    rename = {date_col: "Date", desc_col: "Description", amount_col: "Debit/Credit"}

    # Build per-category DataFrames
    result = {}
    for category in categories:
        mask = final_labels == category
        cat_df = df.loc[mask, [date_col, desc_col, amount_col]].copy()
        cat_df = cat_df.rename(columns=rename)
        cat_df["Match"] = match_col[mask].values
        cat_df["Debit/Credit"] = pd.to_numeric(cat_df["Debit/Credit"], errors="coerce")
        cat_df = cat_df.reset_index(drop=True)
        result[category] = cat_df

        count = len(cat_df)
        total = cat_df["Debit/Credit"].sum()
        display = config["categories"][category].get("display_name", category)
        print(f"  {display}: {count} rows  |  total = {total:+.2f}")

    # Near-miss DataFrame (rows in review band, not already matched)
    review_mask = review_labels_all.notna()
    review_df = df.loc[review_mask, [date_col, desc_col, amount_col]].copy()
    review_df = review_df.rename(columns=rename)
    review_df["Nearest Category"] = review_labels_all[review_mask].apply(
        lambda c: config["categories"].get(c, {}).get("display_name", c)
    ).values
    review_df["Score"] = review_scores_all[review_mask].values
    review_df["Debit/Credit"] = pd.to_numeric(review_df["Debit/Credit"], errors="coerce")
    review_df = review_df.reset_index(drop=True)
    result["_review"] = review_df

    if len(review_df):
        print(f"  Near-miss (review): {len(review_df)} rows  |  score range {review_threshold}–{main_threshold}")

    # Write output
    if output is None:
        p = Path(filepath)
        output = str(p.parent / f"{p.stem}_interest_tds.xlsx")

    _write_output(result, output, config)
    print(f"\nOutput: {output}")

    return result


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Find interest income and TDS/tax transactions in bank statement files."
    )
    parser.add_argument("filepath",             help="Path to .xlsx, .xls, or .csv file")
    parser.add_argument("--locale",    default="india",  help="Config locale: india | us  (default: india)")
    parser.add_argument("--threshold", default=0.55, type=float, help="Cosine similarity threshold (default: 0.55)")
    parser.add_argument("--output",    default=None,     help="Output Excel path (default: <input>_interest_tds.xlsx)")
    args = parser.parse_args()

    find_transactions(args.filepath, args.locale, args.threshold, args.output)
