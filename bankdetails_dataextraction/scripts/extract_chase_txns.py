"""
Chase Business Statement OCR Extractor
Extracts transactions from Chase bank statement images (section-based format).

Sections and sign:
  Deposits and Additions       → positive
  Checks Paid                  → negative
  ATM & Debit Card Withdrawals → negative
  Electronic Withdrawals       → negative
  Service Fees                 → negative

Output CSV: statement_period, date, description, amount, section, source_page

Usage:
    python extract_chase_txns.py <images_folder>
    python extract_chase_txns.py <images_folder> --output result.csv
"""
import sys
import re
import pytesseract
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from PIL import Image
from pathlib import Path

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'


# ── Section detection ───────────────────────────────────────────────────────────

# (keyword, sign, canonical name)  — checked in order, first match wins
_SECTIONS = [
    ('DEPOSIT',          +1, 'Deposits and Additions'),
    ('ADDITION',         +1, 'Deposits and Additions'),
    ('CHECK',            -1, 'Checks Paid'),
    ('ATM',              -1, 'ATM & Debit Card Withdrawals'),
    ('DEBIT CARD',       -1, 'ATM & Debit Card Withdrawals'),
    ('ELECTRONIC',       -1, 'Electronic Withdrawals'),
    ('OTHER WITHDRAWAL', -1, 'Other Withdrawals'),
    ('SERVICE FEE',      -1, 'Service Fees'),
    ('FEES',             -1, 'Fees'),
]

# Stricter keywords used ONLY when matching printed Total lines.
# Prevents ATM sub-totals ("Total ATM Withdrawals & Debits") from matching
# via the loose "ATM" keyword — requires "ATM & DEBIT" instead.
_TOTAL_KEYWORDS = [
    ('DEPOSITS AND ADDITIONS',  +1, 'Deposits and Additions'),
    ('DEPOSITS & ADDITIONS',    +1, 'Deposits and Additions'),
    ('CHECKS PAID',             -1, 'Checks Paid'),
    ('ATM & DEBIT',             -1, 'ATM & Debit Card Withdrawals'),
    ('ELECTRONIC WITHDRAWALS',  -1, 'Electronic Withdrawals'),
    ('OTHER WITHDRAWAL',        -1, 'Other Withdrawals'),
    ('SERVICE FEE',             -1, 'Service Fees'),
    ('FEES',                    -1, 'Fees'),
]

# Sections whose rows we do NOT want as transactions
_SKIP_SECTIONS = ['DAILY ENDING', 'BALANCE', 'SUMMARY', 'CARD TOTALS',
                  'CARD SUMMARY', 'ATM & DEBIT CARD TOTALS',
                  'ATM & DEBIT CARD SUMMARY', 'HOW TO AVOID', 'CUSTOMER SERVICE']

# Row-level skip patterns (totals, headers, footnotes)
_SKIP_ROW_RE = [
    re.compile(r'^\s*total\s+', re.I),
    re.compile(r'^\s*date\s+(description|check)', re.I),
    re.compile(r'^\s*check\s+no', re.I),
    re.compile(r'^\s*if you (see|have|meet)', re.I),
    re.compile(r'^\s*\*+\s*(all|an image)', re.I),
    re.compile(r'^\s*page\s+\d+', re.I),
    re.compile(r'^\s*(beginning|ending)\s+balance', re.I),
    re.compile(r'^\s*instances\s+amount', re.I),
    re.compile(r'^\s*congratulations', re.I),
]

# ACH trace-line patterns — not useful in description
# Covers OCR garbles: 1D: (ID:), Ind 1D: (Ind ID:)
_TRACE_LINE_RE = re.compile(
    r'^(Trn:|Trace#:|Sec:|Eed:|Ind\s+\S+:|CO Entry|Impound\s|'
    r'[1I]D:|Tax\s+Impoun|Invoice\s+Trn)', re.I)

# Descr:/Deser: (OCR garble) carries the actual payment purpose (NACHA "Company
# Entry Description", e.g. PAYROLL, UTIL BILL) and is usually packed on the same
# OCR line as the noise labels above — capture the payload up to the next label.
_DESCR_LINE_RE = re.compile(
    r'(?:CO Entry\s+)?(?:Descr?:|Deser:)\s*(.+?)'
    r'(?=\s+(?:Trn:|Trace#:|Sec:|Eed:|Ind\s+\S+:|CO Entry|[1I]D:|Tax\s+Impoun|Invoice\s+Trn)|$)',
    re.I)

# Printed section total line: "Total Electronic Withdrawals $6,363.99"
_TOTAL_LINE_RE = re.compile(r'^\s*total\s+(.+?)\s+\$?([\d,]+\.\d{2})\s*\S{0,6}$', re.I)

# Orig CO Name extractor: "Orig CO Name:ACME Corp Orig ID:..." → "ACME Corp"
# Handles OCR variants: Orig |D: / Orig 1D: instead of Orig ID:
_ORIG_CO_RE = re.compile(r'Orig\s+CO\s+Name:\s*(.+?)(?:\s+Orig\s+[|1I!]?D:|$)', re.I)

# Amount at end of line (with optional leading $ and optional trailing OCR noise
# ≤15 chars — wide enough to survive sidebar/barcode bleed like "635.25 ——— 43",
# which has an internal space and would break a plain \S{0,6} bound)
_AMT_TAIL_RE = re.compile(r'\$?([\d,]+\.\d{2})[\s\S]{0,15}$')

# Transaction row: MM/DD (optional trailing period from OCR) followed by something.
# Separator may OCR as underscores instead of whitespace (e.g. "01/30___Zelle...",
# likely an underline artifact) — without this a whole transaction row is missed
# and silently merged into the previous row's amount.
_DATE_ROW_RE = re.compile(r'^(\d{2}/\d{2})\.?[\s_]+(.+)$')

# Zelle confirmation codes / deposit reference numbers are printed directly after
# the name/label with just a space, no delimiter (e.g. "Zelle Payment From
# Christina Alfaro Bacqsyy8Ltxf", "Deposit 7680095258"). Strip the trailing token
# when it contains a digit — real names/labels never do — since these are IDs
# that must not reach the tagger.
_ZELLE_OR_DEPOSIT_RE = re.compile(r'^(Zelle Payment (?:From|To)\s+.+|Deposit)\s+([A-Za-z0-9]+)$', re.I)

# Checks Paid row: check_num [**] DATE AMOUNT
# Date field may be OCR-garbled (e.g. "4 O16" instead of "01/16") — match 1-2 tokens before amount.
_CHECK_ROW_RE = re.compile(r'^(\d{3,6})\s+(?:\*+\s+)?(\S+(?:\s+\S+)?)\s+\$?([\d,]+\.\d{2})')

# Statement period: "January 01, 2025 through January 31, 2025"
_PERIOD_RE = re.compile(
    r'(January|February|March|April|May|June|July|August|September|October|November|December)'
    r'\s+\d+,\s*\d{4}\s+through\s+\S+\s+\d+,\s*(\d{4})', re.I)


# ── Helpers ─────────────────────────────────────────────────────────────────────

def _natural_sort_key(f):
    parts = re.split(r'(\d+)', str(f))
    return [int(p) if p.isdigit() else p.lower() for p in parts]


def _parse_amount(s):
    m = _AMT_TAIL_RE.search(s)
    return float(m.group(1).replace(',', '')) if m else None


def _clean_desc(raw):
    """Strip trailing amount and clean up OCR noise / IDs from description."""
    desc = _AMT_TAIL_RE.sub('', raw).strip().rstrip('$').strip()
    # Strip leading OCR noise chars (underscores, dashes, pipes)
    desc = re.sub(r'^[_\-—|~]+\s*', '', desc)
    # Strip a redundant leading date (e.g. "01/10 Payment..." → "Payment...")
    desc = re.sub(r'^\d{2}/\d{2}\.?\s+', '', desc)
    # Collapse Orig CO Name to just the company
    m = _ORIG_CO_RE.search(desc)
    if m:
        desc = 'ACH: ' + m.group(1).strip()
    # Strip explicit transaction/reference IDs — must never reach the tagger
    desc = re.sub(r'\s*Transaction#:\s*\d+', '', desc, flags=re.I)
    desc = re.sub(r'\.\.\.\d{3,6}', '', desc)
    # Strip Zelle/Deposit trailing confirmation code (see _ZELLE_OR_DEPOSIT_RE)
    zm = _ZELLE_OR_DEPOSIT_RE.match(desc)
    if zm and re.search(r'\d', zm.group(2)):
        desc = zm.group(1).strip()
    desc = re.sub(r'\s{2,}', ' ', desc).strip()
    return desc


def _parse_total_line(line):
    """Extract (section_name, amount) from a printed 'Total X $Y' line, or (None, None).
    Uses _TOTAL_KEYWORDS (stricter than _SECTIONS) to avoid matching ATM sub-totals
    like 'Total ATM Withdrawals & Debits' via the loose 'ATM' keyword."""
    m = _TOTAL_LINE_RE.match(line)
    if not m:
        return None, None
    label = m.group(1).strip().upper()
    amt = float(m.group(2).replace(',', ''))
    for keyword, _sign, name in _TOTAL_KEYWORDS:
        if keyword in label:
            return name, amt
    return None, None


def _detect_section(line):
    """
    Return (sign, name) if line is a section header.
    sign=0 means skip section. Returns None if not a header.
    """
    upper = line.upper()
    for keyword in _SKIP_SECTIONS:
        if keyword in upper:
            return 0, keyword
    for keyword, sign, name in _SECTIONS:
        if keyword in upper:
            return sign, name
    return None


def _should_skip_row(line):
    return any(pat.search(line) for pat in _SKIP_ROW_RE)


def _extract_period(text):
    m = _PERIOD_RE.search(text)
    if m:
        # Return the full match
        return m.group(0).strip()
    return None


# ── Core parser ─────────────────────────────────────────────────────────────────

def parse_page(text):
    """
    Parse one page of OCR text.
    Returns (transactions, printed_totals) where printed_totals is
    {section_name: positive_amount} from the statement's own Total lines.
    """
    transactions = []
    printed_totals = {}
    current_sign = None
    current_section = None
    current_txn = None

    def _flush():
        nonlocal current_txn
        if current_txn and current_txn['date'] and current_txn['amount'] is not None:
            transactions.append(current_txn)
        current_txn = None

    for raw in text.split('\n'):
        line = raw.strip()
        if not line:
            continue

        # Lines starting with MM/DD are ALWAYS transactions, never section headers.
        # Check this first to prevent keywords like ATM/DEPOSIT inside descriptions
        # from triggering section detection.
        is_date_prefixed = bool(re.match(r'^\d{2}/\d{2}[.\s_]', line))

        if not is_date_prefixed:
            # Capture printed section totals before any skip/section logic
            sec_name, amt = _parse_total_line(line)
            if sec_name:
                printed_totals[sec_name] = printed_totals.get(sec_name, 0.0) + amt
                continue

            # Section header detection (only for non-transaction lines)
            sec = _detect_section(line)
            if sec is not None:
                _flush()
                current_sign, current_section = sec
                continue

        if current_sign is None or current_sign == 0:
            continue

        if _should_skip_row(line):
            _flush()
            continue

        # Checks Paid: check_num [**] date amount — date may be OCR-garbled
        if current_section == 'Checks Paid':
            m = _CHECK_ROW_RE.match(line)
            if m:
                _flush()
                check_num, raw_date, amt_str = m.group(1), m.group(2), m.group(3)
                # Normalise OCR-garbled date: keep only if it looks like MM/DD
                date = raw_date if re.match(r'^\d{2}/\d{2}$', raw_date) else '??/??'
                current_txn = {
                    'date': date,
                    'description': f'Check #{check_num}',
                    'amount': float(amt_str.replace(',', '')) * current_sign,
                    'section': current_section,
                }
            continue

        # Regular section: MM/DD description [amount]
        m = _DATE_ROW_RE.match(line)
        if m:
            _flush()
            date, rest = m.group(1), m.group(2)
            # OCR occasionally hallucinates a stray duplicate date before the real
            # one (e.g. "01/12 01/11 Online Transfer..."). Two MM/DD tokens never
            # legitimately appear back-to-back, so the second one is the real date.
            dup = _DATE_ROW_RE.match(rest)
            if dup:
                date, rest = dup.group(1), dup.group(2)
            amt = _parse_amount(rest)
            desc = _clean_desc(rest) if amt is not None else rest.strip()
            current_txn = {
                'date': date,
                'description': desc,
                'amount': amt * current_sign if amt is not None else None,
                'section': current_section,
            }
        elif current_txn is not None:
            # Continuation line
            descr_match = _DESCR_LINE_RE.search(line)
            if descr_match:
                payload = descr_match.group(1).strip()
                # OCR sometimes drops the space before "Sec:", gluing it onto the
                # prior word (e.g. "Tax Impounsec:CCD" → "Tax Impoun", "Paymentrecsec:Web"
                # → "Paymentrec") — strip it even without a preceding space.
                payload = re.sub(r'(?i)sec:\w*', '', payload).strip(' _-')
                if payload:
                    current_txn['description'] += ' | ' + payload
            if _TRACE_LINE_RE.match(line) or descr_match:
                continue  # skip ACH trace junk (Descr/Deser payload already captured above)
            amt = _parse_amount(line)
            if amt is not None and current_txn['amount'] is None:
                # Amount was on the next line
                current_txn['amount'] = amt * current_sign
            elif amt is None and len(line) < 120:
                # Meaningful text continuation — append to description
                cleaned = _clean_desc(line) if _ORIG_CO_RE.search(line) else line
                current_txn['description'] += ' | ' + cleaned

    _flush()
    return transactions, printed_totals


# ── OCR ─────────────────────────────────────────────────────────────────────────

def ocr_image(image_path):
    img = Image.open(image_path)
    # PSM 6: uniform block — better for columnar tables
    return pytesseract.image_to_string(img, config='--psm 6')


# ── Constants ────────────────────────────────────────────────────────────────────

_IMAGE_EXTS = {'.jpg', '.jpeg', '.png', '.tif', '.tiff'}

_SECTION_ORDER = [
    'Deposits and Additions',
    'Checks Paid',
    'ATM & Debit Card Withdrawals',
    'Electronic Withdrawals',
    'Other Withdrawals',
    'Service Fees',
    'Fees',
]

_HDR_FILL = PatternFill('solid', fgColor='1F4E79')
_HDR_FONT = Font(color='FFFFFF', bold=True)
_BOLD     = Font(bold=True)
_CENTER   = Alignment(horizontal='center', vertical='center', wrap_text=True)
_FMT_NUM  = '#,##0.00'


# ── Processing helpers ────────────────────────────────────────────────────────────

def _image_files(folder):
    return sorted(
        [f for f in folder.iterdir() if f.suffix.lower() in _IMAGE_EXTS],
        key=_natural_sort_key,
    )


def _section_totals(txns):
    totals = {}
    for t in txns:
        totals[t['section']] = totals.get(t['section'], 0.0) + t['amount']
    return totals


def _section_sign(section_name):
    for _, sign, name in _SECTIONS:
        if name == section_name:
            return sign
    return -1


def _try_exclude_phantom_row(txns, section, printed):
    """BTP-1: if extracted exceeds printed by exactly one row, remove that row."""
    sec_txns = [t for t in txns if t['section'] == section]
    extracted = abs(sum(t['amount'] for t in sec_txns))
    excess = round(extracted - printed, 2)
    if excess <= 0.02:
        return txns, None
    for t in sec_txns:
        if abs(abs(t['amount']) - excess) < 0.02:
            return [x for x in txns if x is not t], t.get('description', '')
    return txns, None


def process_folder(folder):
    """OCR every image in folder, reconcile section totals, return transaction dicts."""
    imgs = _image_files(folder)
    if not imgs:
        return []
    all_txns = []
    printed_totals = {}
    for img_path in imgs:
        print(f'    {img_path.name}...', end=' ', flush=True)
        text = ocr_image(img_path)
        txns, page_totals = parse_page(text)
        for t in txns:
            t['source_page'] = img_path.name
            t['flag'] = ''
        print(f'{len(txns)} txns')
        all_txns.extend(txns)
        for sec, amt in page_totals.items():
            printed_totals[sec] = printed_totals.get(sec, 0.0) + amt

    # Reconcile per section
    for sec, printed in printed_totals.items():
        # BTP-1: remove phantom total row if it accounts for entire excess
        all_txns, excluded = _try_exclude_phantom_row(all_txns, sec, printed)
        if excluded:
            print(f'  [BTP-1] Removed phantom row in {sec}: "{excluded[:60]}"')

        extracted = abs(sum(t['amount'] for t in all_txns if t['section'] == sec))
        gap = round(printed - extracted, 2)

        if abs(gap) > 0.01:
            # BTP-2: insert sentinel row — valid rows stay unflagged
            sign = _section_sign(sec)
            print(f'  ⚠  {sec}: extracted ${extracted:,.2f} | statement ${printed:,.2f} | gap ${gap:,.2f}')
            all_txns.append({
                'date': '',
                'description': '*** MISSING ROWS — check section manually ***',
                'amount': round(gap * sign, 2),
                'section': sec,
                'source_page': '(see above)',
                'flag': f'VERIFY: gap ${gap:,.2f} (extracted ${extracted:,.2f} | statement ${printed:,.2f})',
            })
        else:
            print(f'  ✓  {sec}: ${printed:,.2f}')

    return all_txns


# ── Excel output ──────────────────────────────────────────────────────────────────

def _write_txn_sheet(ws, txns):
    """Write transaction rows to a worksheet."""
    headers = ['Date', 'Description', 'Amount', 'Section', 'Source Page', 'Flag']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment = _HDR_FONT, _HDR_FILL, _CENTER

    red_font = Font(color='FF0000')
    for r, t in enumerate(txns, 2):
        ws.cell(row=r, column=1, value=t['date'])
        ws.cell(row=r, column=2, value=t['description'])
        amt = ws.cell(row=r, column=3, value=t['amount'])
        amt.number_format = _FMT_NUM
        ws.cell(row=r, column=4, value=t['section'])
        ws.cell(row=r, column=5, value=t['source_page'])
        flag_cell = ws.cell(row=r, column=6, value=t.get('flag', ''))
        if t.get('flag'):
            flag_cell.font = red_font

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 52
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 26
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 45
    ws.freeze_panes = 'A2'


def _write_summary_sheet(ws, sheets_data):
    """Write summary tab: one row per month, totals by section + net."""
    sections = [s for s in _SECTION_ORDER
                if any(s in _section_totals(txns) for _, txns in sheets_data)]
    headers = ['Month', 'Transactions'] + sections + ['Net']

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment = _HDR_FONT, _HDR_FILL, _CENTER
    ws.row_dimensions[1].height = 36

    for r, (name, txns) in enumerate(sheets_data, 2):
        totals = _section_totals(txns)
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=len(txns))
        net = 0.0
        for col, sec in enumerate(sections, 3):
            val = round(totals.get(sec, 0.0), 2)
            ws.cell(row=r, column=col, value=val).number_format = _FMT_NUM
            net += val
        ws.cell(row=r, column=len(headers), value=round(net, 2)).number_format = _FMT_NUM

    # Grand total row
    gr = len(sheets_data) + 2
    ws.cell(row=gr, column=1, value='TOTAL').font = _BOLD
    ws.cell(row=gr, column=2, value=sum(len(t) for _, t in sheets_data)).font = _BOLD
    net_total = 0.0
    for col, sec in enumerate(sections, 3):
        val = round(sum(_section_totals(t).get(sec, 0.0) for _, t in sheets_data), 2)
        c = ws.cell(row=gr, column=col, value=val)
        c.number_format, c.font = _FMT_NUM, _BOLD
        net_total += val
    c = ws.cell(row=gr, column=len(headers), value=round(net_total, 2))
    c.number_format, c.font = _FMT_NUM, _BOLD

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 14
    for i in range(len(sections) + 1):
        col_letter = openpyxl.utils.get_column_letter(3 + i)
        ws.column_dimensions[col_letter].width = 28


def write_excel(output_path, sheets_data):
    """sheets_data: list of (tab_name, txns). Summary tab inserted first."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    summary_ws = wb.create_sheet(title='Summary')
    for name, txns in sheets_data:
        _write_txn_sheet(wb.create_sheet(title=name[:31]), txns)

    _write_summary_sheet(summary_ws, sheets_data)
    wb.save(output_path)
    print(f'Written to: {output_path}')


# ── Main ─────────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print('Usage: python extract_chase_txns.py <folder> [--output file.xlsx]')
        sys.exit(1)

    folder = Path(sys.argv[1])
    if not folder.is_dir():
        print(f'Error: {folder} is not a directory')
        sys.exit(1)

    output_file = folder.parent / f'{folder.name}_transactions.xlsx'
    if '--output' in sys.argv:
        idx = sys.argv.index('--output')
        if idx + 1 < len(sys.argv):
            output_file = Path(sys.argv[idx + 1])

    # Subfolder mode: each subfolder = one tab
    subfolders = sorted([f for f in folder.iterdir() if f.is_dir()], key=_natural_sort_key)
    if subfolders:
        sheets_data = []
        for sub in subfolders:
            if not _image_files(sub):
                continue
            print(f'\n{sub.name}:')
            txns = process_folder(sub)
            sheets_data.append((sub.name, txns))
            _print_section_summary(txns)
    else:
        # Flat mode: all images in one tab
        print(f'Processing {folder.name}...')
        txns = process_folder(folder)
        sheets_data = [(folder.name, txns)]
        _print_section_summary(txns)

    total = sum(len(t) for _, t in sheets_data)
    print(f'\nTotal transactions across all months: {total}')
    write_excel(output_file, sheets_data)


def _print_section_summary(txns):
    for sec, total in _section_totals(txns).items():
        print(f'  {sec}: ${total:,.2f}')


if __name__ == '__main__':
    main()
