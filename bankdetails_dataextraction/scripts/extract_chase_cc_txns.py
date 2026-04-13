"""
Chase Credit Card Statement OCR Extractor
Extracts transactions from the ACCOUNT ACTIVITY section of Chase CC statement images.

Output Excel: date, description, amount, source_page, flag

Usage:
    python extract_chase_cc_txns.py <images_folder>
    python extract_chase_cc_txns.py <images_folder> --output file.xlsx
"""
import sys
import re
import pytesseract
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from PIL import Image
from pathlib import Path

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

_IMAGE_EXTS = {'.jpg', '.jpeg', '.png', '.tif', '.tiff'}

_DATE_ROW_RE  = re.compile(r'^(\d{2}/\d{2})\.?\s+(.+)$')
_AMT_TAIL_RE  = re.compile(r'-?\$?([\d,]+\.\d{2})\s*\S{0,6}$')
_SIGNED_AMT_RE = re.compile(r'(-?\$?[\d,]+\.\d{2})\s*\S{0,6}$')

# Statement period: "Opening/Closing Date MM/DD/YY - MM/DD/YY"
_PERIOD_RE = re.compile(r'Opening/Closing Date\s+(\d{2}/\d{2}/\d{2})\s*-\s*(\d{2}/\d{2}/\d{2})', re.I)
_CYCLE_TOTAL_RE = re.compile(r'TRANSACTIONS THIS CYCLE.*\$([0-9,]+\.\d{2})', re.I)

_HDR_FILL = PatternFill('solid', fgColor='1F4E79')
_HDR_FONT = Font(color='FFFFFF', bold=True)
_CENTER   = Alignment(horizontal='center', vertical='center', wrap_text=True)
_FMT_NUM  = '#,##0.00'


def _natural_sort_key(f):
    parts = re.split(r'(\d+)', str(f))
    return [int(p) if p.isdigit() else p.lower() for p in parts]


def _image_files(folder):
    return sorted(
        [f for f in folder.iterdir() if f.suffix.lower() in _IMAGE_EXTS],
        key=_natural_sort_key,
    )


def ocr_image(image_path):
    return pytesseract.image_to_string(Image.open(image_path), config='--psm 6')


def _parse_amount(s):
    """Extract signed amount from end of string."""
    m = re.search(r'(-?[\d,]+\.\d{2})\s*\S{0,6}$', s)
    return float(m.group(1).replace(',', '')) if m else None


def _clean_desc(raw):
    """Strip trailing amount from description."""
    desc = re.sub(r'\s+-?[\d,]+\.\d{2}\s*\S{0,6}$', '', raw).strip()
    desc = re.sub(r'^[_\-—|~]+\s*', '', desc)
    return desc


def _extract_period(text):
    m = _PERIOD_RE.search(text)
    return f"{m.group(1)} - {m.group(2)}" if m else ''


def parse_page(text):
    """Extract transactions from ACCOUNT ACTIVITY section."""
    transactions = []
    printed_total = None
    in_activity = False

    for raw in text.split('\n'):
        line = raw.strip()
        if not line:
            continue

        if 'ACCOUNT ACTIVITY' in line.upper() or 'TRANSACTION MERCHANT NAME' in line.upper():
            in_activity = True
            continue

        if in_activity and 'TRANSACTIONS THIS CYCLE' in line.upper():
            m = _CYCLE_TOTAL_RE.search(line)
            if m:
                printed_total = float(m.group(1).replace(',', ''))
            break

        if not in_activity:
            continue

        m = _DATE_ROW_RE.match(line)
        if m:
            date, rest = m.group(1), m.group(2)
            amt = _parse_amount(rest)
            if amt is not None:
                transactions.append({
                    'date': date,
                    'description': _clean_desc(rest),
                    'amount': amt,
                    'flag': '',
                })

    return transactions, printed_total


def process_folder(folder):
    imgs = _image_files(folder)
    if not imgs:
        return [], ''

    all_txns = []
    printed_total = None
    period = ''
    for img_path in imgs:
        print(f'    {img_path.name}...', end=' ', flush=True)
        text = ocr_image(img_path)
        if not period:
            period = _extract_period(text)
        txns, page_total = parse_page(text)
        for t in txns:
            t['source_page'] = img_path.name
        print(f'{len(txns)} txns')
        all_txns.extend(txns)
        if page_total is not None:
            printed_total = page_total

    # Reconcile against TRANSACTIONS THIS CYCLE total
    if printed_total is not None:
        extracted = round(abs(sum(t['amount'] for t in all_txns)), 2)
        gap = round(printed_total - extracted, 2)
        if abs(gap) > 0.01:
            print(f'  ⚠  Extracted ${extracted:,.2f} | Statement ${printed_total:,.2f} | Gap ${gap:,.2f}')
            all_txns.append({
                'date': '', 'description': '*** MISSING ROWS — check manually ***',
                'amount': gap, 'flag': f'VERIFY: gap ${gap:,.2f}', 'source_page': '(see above)',
            })
        else:
            print(f'  ✓  Reconciled: ${printed_total:,.2f}')
    else:
        print('  ⚠  No TRANSACTIONS THIS CYCLE total found — reconciliation skipped')

    print(f'  Total: {len(all_txns)} transactions')
    return all_txns, period


def write_excel(output_path, sheets_data):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for tab_name, txns in sheets_data:
        ws = wb.create_sheet(title=tab_name[:31])
        headers = ['Date', 'Description', 'Amount', 'Source Page', 'Flag']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font, c.fill, c.alignment = _HDR_FONT, _HDR_FILL, _CENTER

        red_font = Font(color='FF0000')
        for r, t in enumerate(txns, 2):
            ws.cell(row=r, column=1, value=t['date'])
            ws.cell(row=r, column=2, value=t['description'])
            ws.cell(row=r, column=3, value=t['amount']).number_format = _FMT_NUM
            ws.cell(row=r, column=4, value=t.get('source_page', ''))
            flag_cell = ws.cell(row=r, column=5, value=t.get('flag', ''))
            if t.get('flag'):
                flag_cell.font = red_font

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 52
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 45
        ws.freeze_panes = 'A2'

    wb.save(output_path)
    print(f'Written to: {output_path}')


def main():
    if len(sys.argv) < 2:
        print('Usage: python extract_chase_cc_txns.py <folder> [--output file.xlsx]')
        sys.exit(1)

    folder = Path(sys.argv[1])
    if not folder.is_dir():
        print(f'Error: {folder} is not a directory')
        sys.exit(1)

    output_file = folder.parent / f'{folder.name}_cc_transactions.xlsx'
    if '--output' in sys.argv:
        idx = sys.argv.index('--output')
        if idx + 1 < len(sys.argv):
            output_file = Path(sys.argv[idx + 1])

    subfolders = sorted([f for f in folder.iterdir() if f.is_dir()], key=_natural_sort_key)
    if subfolders:
        sheets_data = []
        for sub in subfolders:
            if not _image_files(sub):
                continue
            print(f'\n{sub.name}:')
            txns, period = process_folder(sub)
            tab = period.replace('/', '-') if period else sub.name
            sheets_data.append((tab, txns))
    else:
        print(f'Processing {folder.name}...')
        txns, period = process_folder(folder)
        tab = period.replace('/', '-') if period else folder.name
        sheets_data = [(tab, txns)]

    total = sum(len(t) for _, t in sheets_data)
    print(f'\nTotal transactions: {total}')
    write_excel(output_file, sheets_data)


if __name__ == '__main__':
    main()
