"""
Capital One Bank Statement OCR Extractor
Extracts transactions from Capital One bank statement images.
Outputs one Excel file per account: Summary | Master | one tab per month.

Accounts detected automatically from statement headers.
Amount format: "- $88.29" (withdrawal) / "+ $2,400.00" (deposit)

Usage:
    python extract_capitalone_txns.py <images_folder>
    python extract_capitalone_txns.py <images_folder> --output-dir <path>
"""

import sys
import re
import pytesseract
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from PIL import Image
from pathlib import Path

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'


# ── Month constants ────────────────────────────────────────────────────────────

_MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
           'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
_MONTH_NUM = {m: i + 1 for i, m in enumerate(_MONTHS)}


# ── OCR / parse patterns ───────────────────────────────────────────────────────

# Account section header: "Simply Checking - 2166066502"
_ACCOUNT_HDR_RE = re.compile(
    r'(Simply\s+Checking|Confidence\s+Savings|Performance\s+Savings)\s*[-–—]\s*(\d{7,})',
    re.I)

# Transaction date: "Jan 6", "Jan7" (no space), "Jans" / "Jani" (garbled digit)
_TXN_DATE_RE = re.compile(
    r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]?\s*(\d{0,2})\s+(.+)',
    re.I)

# Signed amount: "- $88.29" or "+ $2,400.00"
_SIGNED_AMT_RE = re.compile(r'([+-])\s*\$\s*([\d,]+\.\d{2})')

# Trailing balance: last "$amount" on a line
_BAL_TAIL_RE = re.compile(r'\$?\s*([\d,]+\.\d{2})\s*$')

_OPENING_RE = re.compile(r'Opening\s+Balance', re.I)
_CLOSING_RE = re.compile(r'Closing\s+Balance', re.I)
_FEES_RE    = re.compile(r'Fees?\s+Summary', re.I)

# Lines to always discard
_SKIP_RES = [
    re.compile(r'^\s*DATE\s+DESCRIPTION', re.I),
    re.compile(r'^\s*JOINT\s+WITH', re.I),
    re.compile(r'^\s*Page\s+\d+', re.I),
    re.compile(r'^\s*capitalone\.com', re.I),
    re.compile(r'^\s*P\.O\.\s*Box', re.I),
    re.compile(r'^\s*STATEMENT\s+PERIOD', re.I),
    re.compile(r'^\s*RAMESH', re.I),
    re.compile(r'^\s*(Thanks\s+for|Here.s\s+your|TOTAL\s+ENDING|IN\s+ALL\s+ACCOUNTS)', re.I),
    re.compile(r'^\s*(Account\s+Summary|Cashflow\s+Summary)', re.I),
    re.compile(r'(Simply\s+Checking|Confidence\s+Savings)\.\.\.', re.I),  # truncated summary table rows
    re.compile(r'^\s*All\s+Accounts', re.I),
    re.compile(r'^\s*(TOTAL\s+FOR\s+THIS|TOTAL\s+YEAR|Total\s+Overdraft|Total\s+Return|Total\s+Fees)', re.I),
    re.compile(r'^\s*PERIOD\s+DATE', re.I),
    re.compile(r'^\s*(ANNUAL|APY|YTD\s+INTEREST|DAYS\s+IN\s+STATEMENT)', re.I),
    re.compile(r'^\s*\d+\.\d{2}\s*%', re.I),      # APY percentage
    re.compile(r'^\s*1-888', re.I),                # phone number
    re.compile(r'^\s*FDIC', re.I),
    re.compile(r'^\s*\$[\d,]+\.\d{2}\s*$', re.I), # standalone large balance (summary page)
    re.compile(r'^\s*CATEGORY\s*$', re.I),
    re.compile(r'^\s*(AMOUNT|BALANCE)\s*$', re.I),
]


# ── Excel styling ──────────────────────────────────────────────────────────────

_HDR_FILL = PatternFill('solid', fgColor='1F4E79')
_HDR_FONT = Font(color='FFFFFF', bold=True)
_BOLD     = Font(bold=True)
_CENTER   = Alignment(horizontal='center', vertical='center', wrap_text=True)
_FMT_NUM  = '#,##0.00'
_RED_FONT = Font(color='FF0000')


# ── Helpers ────────────────────────────────────────────────────────────────────

def _natural_sort_key(path):
    parts = re.split(r'(\d+)', str(path))
    return [int(p) if p.isdigit() else p.lower() for p in parts]


def _month_from_filename(path):
    """Return ('Jan 2025', 2025) from '20250101-Bank statement-1.jpg'."""
    m = re.match(r'(\d{4})(\d{2})\d{2}', Path(path).name)
    if m:
        year, month = int(m.group(1)), int(m.group(2))
        return f"{_MONTHS[month - 1]} {year}", year
    return None, None


def _period_sort_key(period):
    parts = period.split()
    if len(parts) == 2:
        try:
            return (int(parts[1]), _MONTH_NUM.get(parts[0], 0))
        except ValueError:
            pass
    return (9999, 0)


def _to_float(val):
    if not val:
        return None
    try:
        return float(str(val).replace(',', '').replace('$', '').strip())
    except (ValueError, AttributeError):
        return None


def _should_skip(line):
    return any(pat.search(line) for pat in _SKIP_RES)


# ── OCR ────────────────────────────────────────────────────────────────────────

def _count_txn_lines(text):
    """Count Capital One transaction lines: month prefix + signed amount on same line."""
    date_re = re.compile(
        r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]?\s*\d',
        re.MULTILINE | re.I)
    amt_re = re.compile(r'[+-]\s*\$')
    count = 0
    for m in date_re.finditer(text):
        line_end = text.find('\n', m.start())
        line = text[m.start(): line_end if line_end != -1 else len(text)]
        if amt_re.search(line):
            count += 1
    return count


def ocr_image(image_path):
    img = Image.open(image_path)
    t6 = pytesseract.image_to_string(img, config='--psm 6')
    t3 = pytesseract.image_to_string(img, config='--psm 3')
    return t3 if _count_txn_lines(t3) > _count_txn_lines(t6) else t6


# ── Parser ─────────────────────────────────────────────────────────────────────

def parse_month(images, year):
    """
    Parse all images for one month.
    Returns {last4: {'name': str, 'txns': list, 'opening': float|None, 'closing': float|None}}
    """
    account_data = {}
    current_last4 = None
    in_fees = False
    pending_txn = None

    def _flush():
        nonlocal pending_txn
        if pending_txn is not None and current_last4:
            account_data[current_last4]['txns'].append(pending_txn)
        pending_txn = None

    for img_path in images:
        text = ocr_image(img_path)

        for raw in text.split('\n'):
            line = raw.strip()
            if not line:
                continue

            # Account header — highest priority, resets fees section
            m_acct = _ACCOUNT_HDR_RE.search(line)
            if m_acct:
                _flush()
                acct_type = re.sub(r'\s+', ' ', m_acct.group(1).strip())
                last4 = m_acct.group(2).strip()[-4:]
                current_last4 = last4
                in_fees = False
                if last4 not in account_data:
                    account_data[last4] = {
                        'name':    f"{acct_type} {last4}",
                        'txns':    [],
                        'opening': None,
                        'closing': None,
                    }
                continue

            # Fees Summary section start
            if _FEES_RE.search(line):
                _flush()
                in_fees = True
                continue

            if in_fees:
                continue

            if _should_skip(line):
                continue

            if current_last4 is None:
                continue

            # Opening Balance
            if _OPENING_RE.search(line):
                _flush()
                m_bal = _BAL_TAIL_RE.search(line)
                if m_bal and account_data[current_last4]['opening'] is None:
                    account_data[current_last4]['opening'] = _to_float(m_bal.group(1))
                continue

            # Closing Balance
            if _CLOSING_RE.search(line):
                _flush()
                m_bal = _BAL_TAIL_RE.search(line)
                if m_bal:
                    account_data[current_last4]['closing'] = _to_float(m_bal.group(1))
                continue

            # Transaction date line
            m_date = _TXN_DATE_RE.match(line)
            if m_date:
                rest = m_date.group(3)
                m_amt = _SIGNED_AMT_RE.search(rest)
                if m_amt:
                    _flush()
                    month_abbr = m_date.group(1).capitalize()[:3]
                    day = m_date.group(2).strip()
                    date_str = f"{month_abbr} {day}" if day else f"{month_abbr} ??"

                    sign    = m_amt.group(1)
                    amt_str = m_amt.group(2)
                    desc    = rest[:m_amt.start()].strip()

                    bal_part = rest[m_amt.end():]
                    m_bal    = re.search(r'\$?\s*([\d,]+\.\d{2})\s*$', bal_part)
                    balance  = m_bal.group(1) if m_bal else ''

                    pending_txn = {
                        'date':        date_str,
                        'description': desc,
                        'subtracted':  amt_str if sign == '-' else '',
                        'added':       amt_str if sign == '+' else '',
                        'balance':     balance,
                        'flag':        '',
                    }
                else:
                    # Date prefix but no signed amount — description continuation with garbled date
                    if pending_txn is not None:
                        pending_txn['description'] += ' ' + line
            else:
                # Continuation line
                if pending_txn is not None:
                    pending_txn['description'] += ' | ' + line

        # End of page — flush and reset fees flag
        _flush()
        in_fees = False

    return account_data


# ── Reconciliation ─────────────────────────────────────────────────────────────

def _reconcile(txns):
    """Balance walk. Flags VERIFY on mismatches. Returns flag count."""
    flags = 0
    prev_bal = None
    for t in txns:
        bal = _to_float(t['balance'])
        sub = _to_float(t['subtracted'])
        add = _to_float(t['added'])
        if bal is None:
            continue
        if prev_bal is None:
            prev_bal = bal
            continue
        expected = round(prev_bal + (add or 0) - (sub or 0), 2)
        if abs(expected - bal) > 0.02:
            note = f'VERIFY: expected {expected:.2f}, got {bal:.2f}'
            t['flag'] = (t['flag'] + ' | ' + note).lstrip(' |').strip()
            flags += 1
        prev_bal = bal
    return flags


# ── Excel output ───────────────────────────────────────────────────────────────

def _write_txn_sheet(ws, txns):
    headers = ['Date', 'Description', 'Subtracted', 'Added', 'Balance', 'Flag']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment = _HDR_FONT, _HDR_FILL, _CENTER

    for r, t in enumerate(txns, 2):
        ws.cell(row=r, column=1, value=t['date'])
        ws.cell(row=r, column=2, value=t['description'])
        for col, field in [(3, 'subtracted'), (4, 'added'), (5, 'balance')]:
            v = _to_float(t.get(field, ''))
            c = ws.cell(row=r, column=col, value=v)
            if v is not None:
                c.number_format = _FMT_NUM
        flag = t.get('flag', '')
        fc = ws.cell(row=r, column=6, value=flag)
        if flag:
            fc.font = _RED_FONT

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 45
    ws.freeze_panes = 'A2'


def _write_master_sheet(ws, sheets_data):
    headers = ['Month', 'Date', 'Description', 'Subtracted', 'Added', 'Balance', 'Flag']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment = _HDR_FONT, _HDR_FILL, _CENTER

    r = 2
    for period, txns in sheets_data:
        for t in txns:
            ws.cell(row=r, column=1, value=period)
            ws.cell(row=r, column=2, value=t['date'])
            ws.cell(row=r, column=3, value=t['description'])
            for col, field in [(4, 'subtracted'), (5, 'added'), (6, 'balance')]:
                v = _to_float(t.get(field, ''))
                c = ws.cell(row=r, column=col, value=v)
                if v is not None:
                    c.number_format = _FMT_NUM
            flag = t.get('flag', '')
            fc = ws.cell(row=r, column=7, value=flag)
            if flag:
                fc.font = _RED_FONT
            r += 1

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 45
    ws.freeze_panes = 'A2'


def _write_summary_sheet(ws, sheets_data, month_meta):
    """
    month_meta: {period: {'opening': float|None, 'closing': float|None}}
    Columns: Month | Txns | Subtracted | Added | Net | Opening | Closing | Calc Closing | Gap
    Gap = Calc Closing - Actual Closing (red if non-zero)
    """
    headers = ['Month', 'Transactions', 'Subtracted', 'Added', 'Net',
               'Opening Balance', 'Closing Balance', 'Calc Closing', 'Gap']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment = _HDR_FONT, _HDR_FILL, _CENTER
    ws.row_dimensions[1].height = 30

    total_txns = 0
    total_sub = total_add = 0.0

    for r, (period, txns) in enumerate(sheets_data, 2):
        sub = round(sum(_to_float(t['subtracted']) or 0 for t in txns), 2)
        add = round(sum(_to_float(t['added']) or 0 for t in txns), 2)
        net = round(add - sub, 2)
        meta     = month_meta.get(period, {})
        opening  = meta.get('opening')
        closing  = meta.get('closing')
        calc_cls = round(opening + add - sub, 2) if opening is not None else None
        gap      = round(calc_cls - closing, 2) if (calc_cls is not None and closing is not None) else None

        ws.cell(row=r, column=1, value=period)
        ws.cell(row=r, column=2, value=len(txns))
        ws.cell(row=r, column=3, value=sub).number_format  = _FMT_NUM
        ws.cell(row=r, column=4, value=add).number_format  = _FMT_NUM
        ws.cell(row=r, column=5, value=net).number_format  = _FMT_NUM
        for col, val in [(6, opening), (7, closing), (8, calc_cls)]:
            c = ws.cell(row=r, column=col)
            if val is not None:
                c.value, c.number_format = val, _FMT_NUM
            else:
                c.value = 'N/A'
        c = ws.cell(row=r, column=9)
        if gap is not None:
            c.value, c.number_format = gap, _FMT_NUM
            if abs(gap) > 0.02:
                c.font = _RED_FONT
        else:
            c.value = 'N/A'

        total_txns += len(txns)
        total_sub  += sub
        total_add  += add

    gr = len(sheets_data) + 2
    ws.cell(row=gr, column=1, value='TOTAL').font = _BOLD
    ws.cell(row=gr, column=2, value=total_txns).font = _BOLD
    for col, val in [(3, total_sub), (4, total_add), (5, total_add - total_sub)]:
        c = ws.cell(row=gr, column=col, value=round(val, 2))
        c.number_format, c.font = _FMT_NUM, _BOLD

    for col_letter, width in zip('ABCDEFGHI', [10, 14, 14, 14, 12, 16, 16, 16, 12]):
        ws.column_dimensions[col_letter].width = width


def write_excel(output_path, sheets_data, month_meta):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    summary_ws = wb.create_sheet(title='Summary')
    master_ws  = wb.create_sheet(title='Master')
    for period, txns in sheets_data:
        _write_txn_sheet(wb.create_sheet(title=period), txns)
    _write_summary_sheet(summary_ws, sheets_data, month_meta)
    _write_master_sheet(master_ws, sheets_data)
    wb.save(output_path)
    print(f'  Written: {output_path}')


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print('Usage: python extract_capitalone_txns.py <folder> [--output-dir <dir>]')
        sys.exit(1)

    folder = Path(sys.argv[1])
    if not folder.is_dir():
        print(f'Error: {folder} is not a directory')
        sys.exit(1)

    output_dir = folder.parent
    if '--output-dir' in sys.argv:
        idx = sys.argv.index('--output-dir')
        if idx + 1 < len(sys.argv):
            output_dir = Path(sys.argv[idx + 1])
    output_dir.mkdir(parents=True, exist_ok=True)

    images = sorted(
        [f for f in folder.iterdir() if f.suffix.lower() in ('.jpg', '.jpeg', '.png', '.tif', '.tiff')],
        key=_natural_sort_key,
    )
    print(f'Found {len(images)} images in {folder}')

    # Group images by YYYYMMDD filename prefix → one group per month
    month_groups: dict = {}
    for img in images:
        period, year = _month_from_filename(img)
        if period:
            month_groups.setdefault((period, year), []).append(img)

    # Parse all months; accumulate per-account data
    # all_accounts: {last4: {'name': str, 'months': {period: {txns, opening, closing}}}}
    all_accounts: dict = {}

    for (period, year), imgs in sorted(month_groups.items(), key=lambda x: _period_sort_key(x[0][0])):
        print(f'\n{period}: ({len(imgs)} pages)')
        acct_data = parse_month(imgs, year)

        for last4, data in acct_data.items():
            if last4 not in all_accounts:
                all_accounts[last4] = {'name': data['name'], 'months': {}}
            all_accounts[last4]['months'][period] = {
                'txns':    data['txns'],
                'opening': data['opening'],
                'closing': data['closing'],
            }
            flags = _reconcile(data['txns'])
            o_str = f"${data['opening']:,.2f}" if data['opening'] is not None else 'N/A'
            c_str = f"${data['closing']:,.2f}" if data['closing'] is not None else 'N/A'
            flag_str = f' | {flags} VERIFY' if flags else ''
            print(f'  {data["name"]}: {len(data["txns"])} txns  '
                  f'Opening {o_str} → Closing {c_str}{flag_str}')

    # Write one Excel file per account
    print()
    for last4, acct in sorted(all_accounts.items()):
        name = acct['name']
        sheets_data = sorted(
            [(p, m['txns']) for p, m in acct['months'].items()],
            key=lambda x: _period_sort_key(x[0]),
        )
        month_meta = {
            p: {'opening': acct['months'][p]['opening'],
                'closing': acct['months'][p]['closing']}
            for p in acct['months']
        }
        safe = re.sub(r'\s+', '_', name).lower()
        write_excel(output_dir / f'capitalone_{safe}_transactions.xlsx',
                    sheets_data, month_meta)

    print('\nDone.')


if __name__ == '__main__':
    main()
