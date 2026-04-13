"""
Indian Bank Statement PDF Extractor
Extracts transactions from Indian bank statement PDFs.
Outputs Excel with:
  Sheet 1 — Transactions: Date, Description, Amount (signed +/-)
  Sheet 2 — Reconciliation: opening/closing balance check + row-by-row gap detection

Supported formats (auto-detected from PDF content):
  statement_report  — HDFC-like 5-col table PDFs
  kotak_bank        — Kotak Bank text-based PDFs
  hdfc_fd           — HDFC FD/Savings scanned/rotated PDFs (OCR via PyMuPDF + Tesseract)

Usage:
    python extract_india_bank_txns.py <file.pdf>
    python extract_india_bank_txns.py <folder/>           # processes all PDFs in folder
    python extract_india_bank_txns.py <file.pdf> --output result.xlsx
"""

import sys
import re
import io
import yaml
from pathlib import Path
from datetime import datetime

import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# OCR imports — only required for hdfc_fd (scanned/rotated) format
try:
    import fitz          # PyMuPDF
    import pytesseract
    from PIL import Image
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

CONFIG_FILE = Path(__file__).parent.parent / 'configs' / 'india_bank_configs.yaml'


# ── Config ──────────────────────────────────────────────────────────────────

def load_configs():
    with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)


# ── Shared helpers ───────────────────────────────────────────────────────────

def normalize_date(date_str, fmt):
    """Parse date_str with fmt, return YYYY-MM-DD. Returns None on failure."""
    try:
        return datetime.strptime(date_str.strip(), fmt).strftime('%Y-%m-%d')
    except ValueError:
        return None


def parse_indian_number(s):
    """Parse Indian number format (e.g. 1,23,456.78) to float. Returns None if blank/invalid."""
    if not s or not str(s).strip():
        return None
    cleaned = str(s).strip().replace(',', '')
    try:
        return float(cleaned)
    except ValueError:
        return None


# ── Bank auto-detection ──────────────────────────────────────────────────────

def detect_bank(pdf_path, configs):
    """
    Inspect PDF content and return the matching config key, or None if unknown.
    Order: table-based → text-based → OCR fallback (0 chars).
    """
    with pdfplumber.open(pdf_path) as pdf:
        first = pdf.pages[0]

        # 1. Table-based detection: check header row of first table
        tables = first.extract_tables()
        if tables and tables[0]:
            header_text = ' '.join(str(c) for c in tables[0][0] if c).upper()
            for key, cfg in configs.items():
                if cfg.get('extraction_mode') == 'table':
                    keywords = cfg.get('detect_table_keywords', [])
                    if keywords and all(kw.upper() in header_text for kw in keywords):
                        return key

        # 2. Text-based detection: keyword scan on raw text
        text = first.extract_text() or ''
        if text.strip():
            for key, cfg in configs.items():
                if cfg.get('extraction_mode') == 'text':
                    keywords = cfg.get('detect_text_keywords', [])
                    if keywords and all(kw.upper() in text.upper() for kw in keywords):
                        return key

        # 3. OCR fallback: no extractable text → scanned/image PDF
        if not text.strip():
            for key, cfg in configs.items():
                if cfg.get('extraction_mode') == 'ocr':
                    return key

    return None


# ── Extractor: table mode (statement_report) ─────────────────────────────────

_OPENING_RE = re.compile(r'Opening\s*:\s*([\d,]+\.\d{2})', re.IGNORECASE)
_CLOSING_RE = re.compile(r'Closing Balance\s*:\s*([\d,]+\.\d{2})', re.IGNORECASE)

def extract_statement_report(pdf_path, cfg):
    """
    Extract from PDFs where pdfplumber returns a clean table.
    Columns: Date | Description | Withdrawal | Deposit | Closing Balance
    Sign: Withdrawal → negative, Deposit → positive.
    Also captures opening/closing balance from page header text.
    Returns (rows, opening_balance, closing_balance).
    """
    rows = []
    date_fmt   = cfg['date_format']
    date_col   = cfg['date_col']
    desc_col   = cfg['description_col']
    wdl_col    = cfg['withdrawal_col']
    dep_col    = cfg['deposit_col']
    bal_col    = dep_col + 1          # Closing Balance is the column after Deposit
    skip_kw    = cfg.get('header_skip_keyword', '').lower()
    opening    = None
    closing    = None

    with pdfplumber.open(pdf_path) as pdf:
        # Extract opening/closing from first-page text
        first_text = pdf.pages[0].extract_text() or ''
        m = _OPENING_RE.search(first_text)
        if m:
            opening = parse_indian_number(m.group(1))
        m = _CLOSING_RE.search(first_text)
        if m:
            closing = parse_indian_number(m.group(1))

        for page in pdf.pages:
            for table in page.extract_tables():
                for row in table:
                    if not row or not row[date_col]:
                        continue
                    date_val = str(row[date_col]).strip()
                    if skip_kw and date_val.lower() == skip_kw:
                        continue
                    date_norm = normalize_date(date_val, date_fmt)
                    if not date_norm:
                        continue

                    desc       = str(row[desc_col] or '').replace('\n', ' ').strip()
                    withdrawal = parse_indian_number(row[wdl_col])
                    deposit    = parse_indian_number(row[dep_col])
                    balance    = parse_indian_number(row[bal_col]) if bal_col < len(row) else None

                    if withdrawal is not None and withdrawal > 0:
                        amount = -withdrawal
                    elif deposit is not None and deposit > 0:
                        amount = deposit
                    else:
                        continue  # null-not-guess

                    rows.append({
                        'date': date_norm, 'description': desc,
                        'amount': amount, '_balance': balance,
                    })

    # The PDF "Opening" field = closing balance of the OLDEST transaction (not before it).
    # Compute the true mathematical opening from the last extracted row (rows are descending).
    if rows and rows[-1]['_balance'] is not None:
        opening = rows[-1]['_balance'] - rows[-1]['amount']

    return rows, opening, closing


# ── Extractor: text mode (kotak_bank) ────────────────────────────────────────

# Kotak has two row types — both matched by a single relaxed pattern:
#   Standard:  N DATE DATE DESCRIPTION REF AMOUNT BALANCE   (debit, ref is last desc word)
#   Credit:    N DATE DATE DESCRIPTION +AMOUNT BALANCE      (IMPS/interest, no trailing ref)
# Ref no is stripped as post-processing (last word if 10+ chars alphanumeric with digit).
_KOTAK_ROW = re.compile(
    r'^\d+\s+'                           # row number (discarded)
    r'(\d{2}\s+\w{3}\s+\d{4})\s+'       # transaction date: DD Mon YYYY
    r'\d{2}\s+\w{3}\s+\d{4}\s+'         # value date (discarded)
    r'(.+?)\s+'                          # description (lazy — stops before amount)
    r'(\+?-?[0-9,]+\.[0-9]{2})\s+'      # signed amount (+ = credit, - = debit)
    r'([0-9,]+\.[0-9]{2})'              # closing balance
    r'\s*$'
)

def _strip_kotak_ref(desc):
    """Remove trailing ref-no token from description (10+ chars, alphanumeric, has digit)."""
    parts = desc.rsplit(None, 1)
    if len(parts) == 2 and re.match(r'^[A-Z0-9]{10,}$', parts[1]) and re.search(r'\d', parts[1]):
        return parts[0].strip()
    return desc

def extract_kotak(pdf_path, cfg):
    """
    Extract from Kotak Bank PDFs where table header is found but data rows
    are only in raw text. Amount column is pre-signed (negative = debit, + = credit).
    Returns (rows, opening_balance, closing_balance).
    """
    rows = []
    date_fmt = cfg['date_format']

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ''
            for line in text.splitlines():
                m = _KOTAK_ROW.match(line.strip())
                if not m:
                    continue
                date_str, desc, amount_str, balance_str = m.groups()
                date_str  = re.sub(r'\s+', ' ', date_str.strip())
                date_norm = normalize_date(date_str, date_fmt)
                if not date_norm:
                    continue
                amount_str = amount_str.lstrip('+')   # strip leading + on credit amounts
                amount  = parse_indian_number(amount_str)
                balance = parse_indian_number(balance_str)
                if amount is None:
                    continue
                desc = _strip_kotak_ref(desc.strip())
                rows.append({
                    'date': date_norm, 'description': desc,
                    'amount': amount, '_balance': balance,
                })

    if not rows:
        return rows, None, None
    # Opening = balance of first row minus its amount (balance before that transaction)
    opening = rows[0]['_balance'] - rows[0]['amount'] if rows[0]['_balance'] is not None else None
    closing = rows[-1]['_balance']
    return rows, opening, closing


# ── Extractor: OCR mode (hdfc_fd) ────────────────────────────────────────────

def _render_best_rotation(page_obj, dpi):
    """
    Render a PyMuPDF page at dpi, trying 0/90/180/270 degree pre-rotations.
    Returns the PIL Image with the most date patterns (best orientation).
    """
    date_re = re.compile(r'\d{2}/\d{2}/\d{2,4}')
    best_img, best_count = None, -1
    for deg in [0, 90, 180, 270]:
        mat = fitz.Matrix(dpi / 72, dpi / 72).prerotate(deg)
        pix = page_obj.get_pixmap(matrix=mat)
        img = Image.open(io.BytesIO(pix.tobytes('png')))
        text = pytesseract.image_to_string(img, config='--psm 6')
        count = len(date_re.findall(text))
        if count > best_count:
            best_count, best_img = count, img
    return best_img


def _extract_amount_from_tokens(tokens):
    """
    Parse an amount from one or two numeric tokens in an amount column.
    Handles:
      - Standard:    "421,220.00" → 421220.00
      - OCR-garbled: "048"        → 0.48  (decimal point dropped, last 2 = decimals)
      - OCR-split:   ["248","04"] → 248.04 (integer + 2-digit decimal on separate tokens)
    Returns float or None (null-not-guess).
    """
    nums = [t for t in tokens if re.match(r'^-?[\d,]+\.?\d*$', t)]
    if not nums:
        return None

    if len(nums) >= 2:
        if re.match(r'^\d+$', nums[0]) and re.match(r'^\d{2}$', nums[1]):
            try:
                return float(nums[0] + '.' + nums[1])
            except ValueError:
                pass

    t = nums[0]
    if re.match(r'^-?\d[\d,]*\.\d{2}$', t):
        return parse_indian_number(t)
    # OCR-garbled: all-digit token, 2–5 chars → insert decimal before last 2
    if re.match(r'^\d{2,5}$', t):
        return float(t[:-2] + '.' + t[-2:])

    return None


def _parse_ocr_page(img, cfg):
    """
    Extract transactions from a single rendered page image using word bounding boxes.
    Sign is determined by x-column position (Withdrawal vs Deposit column).
    Also captures Closing Balance per row for reconciliation.

    Column x boundaries (as fraction of image width, from config):
      [0 ... withdrawal_x_min)                — description / narration zone
      [withdrawal_x_min ... withdrawal_x_max) — Withdrawal Amt  → debit (negative)
      [withdrawal_x_max ... deposit_x_max)    — Deposit Amt     → credit (positive)
      [deposit_x_max ... 1.0)                 — Closing Balance (captured for recon)
    """
    wdl_x_min  = cfg.get('withdrawal_x_min', 0.65)
    wdl_x_max  = cfg.get('withdrawal_x_max', 0.76)
    dep_x_max  = cfg.get('deposit_x_max',   0.91)
    date_fmt   = cfg['date_format']
    date_re    = re.compile(r'^\d{2}/\d{2}/\d{2}$')
    img_width  = img.size[0]

    raw = pytesseract.image_to_data(img, config='--psm 6', output_type=pytesseract.Output.DICT)

    # Group words into y-bands (±20 px tolerance = same row)
    bands = {}
    for i, word in enumerate(raw['text']):
        if not word.strip() or int(raw['conf'][i]) < 20:
            continue
        y = raw['top'][i]
        band_y = next((b for b in bands if abs(b - y) <= 20), y)
        bands.setdefault(band_y, []).append({
            'text': word.strip(),
            'x':    raw['left'][i],
        })

    rows = []
    for y in sorted(bands):
        words = sorted(bands[y], key=lambda w: w['x'])

        # Row must open with a date
        date_words = [w for w in words if date_re.match(w['text'])]
        if not date_words:
            continue
        date_norm = normalize_date(date_words[0]['text'], date_fmt)
        if not date_norm:
            continue

        desc_tokens = []
        wdl_tokens  = []
        dep_tokens  = []
        bal_tokens  = []

        for w in words:
            x_norm = w['x'] / img_width
            text   = w['text']

            if x_norm >= dep_x_max:
                bal_tokens.append(text)    # Closing Balance column
            elif x_norm >= wdl_x_max:
                dep_tokens.append(text)
            elif x_norm >= wdl_x_min:
                wdl_tokens.append(text)
            else:
                # Description zone
                if date_re.match(text):
                    continue  # skip Value Date (second date in row)
                # Strip leading OCR pipe/bracket artifacts (NOT i/I — those are real letters)
                text = re.sub(r'^[|\[\({]+', '', text).strip()
                if not text or len(text) == 1:
                    continue  # drop empty and single-char pipe artifacts
                text_clean = text.rstrip('|,.)([]:;')
                if not text_clean:
                    continue
                # Strip ref-no tokens: 10+ chars, uppercase+digits, must contain a digit
                if re.match(r'^[A-Z0-9]{10,}$', text_clean) and re.search(r'\d', text_clean):
                    continue
                desc_tokens.append(text_clean)

        withdrawal = _extract_amount_from_tokens(wdl_tokens)
        deposit    = _extract_amount_from_tokens(dep_tokens)
        balance    = _extract_amount_from_tokens(bal_tokens)

        if withdrawal is not None:
            amount = -withdrawal
        elif deposit is not None:
            amount = deposit
        else:
            continue  # null-not-guess

        desc = ' '.join(desc_tokens).strip()
        if not desc:
            continue

        rows.append({
            'date': date_norm, 'description': desc,
            'amount': amount, '_balance': balance,
        })

    return rows


def extract_hdfc_fd(pdf_path, cfg):
    if not OCR_AVAILABLE:
        raise RuntimeError(
            "PyMuPDF / pytesseract not installed — cannot process scanned PDFs.\n"
            "Install with: pip install pymupdf pytesseract pillow"
        )
    rows = []
    doc  = fitz.open(str(pdf_path))
    dpi  = cfg.get('ocr_dpi', 300)

    for page_obj in doc:
        img = _render_best_rotation(page_obj, dpi) if cfg.get('ocr_auto_rotate', True) \
              else Image.open(io.BytesIO(
                    page_obj.get_pixmap(matrix=fitz.Matrix(dpi/72, dpi/72)).tobytes('png')
                  ))
        rows.extend(_parse_ocr_page(img, cfg))

    doc.close()

    if not rows:
        return rows, None, None
    opening = rows[0]['_balance'] - rows[0]['amount'] if rows[0]['_balance'] is not None else None
    closing = rows[-1]['_balance']
    return rows, opening, closing


# ── Reconciliation ────────────────────────────────────────────────────────────

def compute_recon(rows, opening, closing_stated, source_name):
    """
    Compute reconciliation for a single PDF.

    Summary check:  opening + sum(amounts) ≈ closing_stated
    Row-by-row check: for consecutive rows, the closing balance of row[i] minus
                      its amount should equal the closing balance of the adjacent row
                      (works for both ascending and descending date order).

    Returns a recon dict.
    """
    rows_with_bal = [r for r in rows if r.get('_balance') is not None]

    total_credits = sum(r['amount'] for r in rows if r['amount'] > 0)
    total_debits  = sum(r['amount'] for r in rows if r['amount'] < 0)
    net           = total_credits + total_debits  # debits are already negative

    computed_closing = (opening + net) if opening is not None else None
    summary_gap      = (computed_closing - closing_stated) \
                       if (computed_closing is not None and closing_stated is not None) else None

    if summary_gap is None:
        summary_status = 'N/A — opening or closing balance not found in PDF'
    elif abs(summary_gap) < 0.02:
        summary_status = '✓ MATCH'
    else:
        summary_status = f'✗ GAP  ₹{summary_gap:+,.2f}  (check for missing transactions)'

    # Row-by-row continuity check
    row_gaps = []
    for i in range(1, len(rows_with_bal)):
        prev = rows_with_bal[i - 1]
        curr = rows_with_bal[i]
        # Implied balance before curr (assuming curr follows prev in time — ascending)
        gap_ascending  = abs((prev['_balance'] + curr['amount']) - curr['_balance'])
        # Implied balance before prev (assuming prev follows curr in time — descending)
        gap_descending = abs((curr['_balance'] + prev['amount']) - prev['_balance'])
        gap = min(gap_ascending, gap_descending)
        if gap > 0.02:
            row_gaps.append({
                'after_row':      i,
                'prev_date':      prev['date'],
                'prev_desc':      prev['description'],
                'prev_amount':    prev['amount'],
                'prev_balance':   prev['_balance'],
                'curr_date':      curr['date'],
                'curr_desc':      curr['description'],
                'curr_amount':    curr['amount'],
                'curr_balance':   curr['_balance'],
                'gap':            gap,
            })

    return {
        'source':           source_name,
        'txn_count':        len(rows),
        'opening':          opening,
        'closing_stated':   closing_stated,
        'total_credits':    total_credits,
        'total_debits':     total_debits,
        'net':              net,
        'computed_closing': computed_closing,
        'summary_gap':      summary_gap,
        'summary_status':   summary_status,
        'row_gaps':         row_gaps,
    }


# ── Dispatch table ────────────────────────────────────────────────────────────

_EXTRACTORS = {
    'table': extract_statement_report,
    'text':  extract_kotak,
    'ocr':   extract_hdfc_fd,
}


def process_pdf(pdf_path, configs):
    bank_key = detect_bank(pdf_path, configs)
    if not bank_key:
        print(f"  WARNING: bank format not recognised — skipping {pdf_path.name}")
        return [], None, None, None
    cfg  = configs[bank_key]
    mode = cfg['extraction_mode']
    fn   = _EXTRACTORS.get(mode)
    if not fn:
        print(f"  WARNING: unknown extraction_mode '{mode}' in config '{bank_key}'")
        return [], None, None, bank_key
    print(f"  Detected: {cfg.get('display_name', bank_key)}  [{mode} mode]")
    rows, opening, closing = fn(pdf_path, cfg)
    return rows, opening, closing, bank_key


# ── Excel output ──────────────────────────────────────────────────────────────

_HDR_FILL  = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
_HDR_FONT  = Font(bold=True, color='FFFFFF')
_DBT_FILL  = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
_CDT_FILL  = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
_GAP_FILL  = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')  # yellow
_ERR_FILL  = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')  # red
_OK_FILL   = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')  # green
_BOLD      = Font(bold=True)
_CENTER    = Alignment(horizontal='center')
_RIGHT     = Alignment(horizontal='right')


def _hdr_row(ws, values):
    ws.append(values)
    for cell in ws[ws.max_row]:
        cell.font      = _HDR_FONT
        cell.fill      = _HDR_FILL
        cell.alignment = _CENTER


def write_excel(all_rows, recon_list, output_path):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Transactions ────────────────────────────────────────────────
    ws_txn = wb.active
    ws_txn.title = 'Transactions'
    _hdr_row(ws_txn, ['Date', 'Description', 'Amount'])

    for row in all_rows:
        ws_txn.append([row['date'], row['description'], row['amount']])
        r = ws_txn.max_row
        fill = _DBT_FILL if row['amount'] < 0 else _CDT_FILL
        for cell in ws_txn[r]:
            cell.fill = fill
        ws_txn.cell(r, 3).number_format = '#,##0.00'

    ws_txn.column_dimensions['A'].width = 14
    ws_txn.column_dimensions['B'].width = 65
    ws_txn.column_dimensions['C'].width = 16

    # ── Sheet 2: Reconciliation ──────────────────────────────────────────────
    ws_rec = wb.create_sheet('Reconciliation')

    for recon in recon_list:
        # Source header
        ws_rec.append([recon['source']])
        ws_rec[ws_rec.max_row][0].font = Font(bold=True, size=12)
        ws_rec.append([])

        # Summary table
        _hdr_row(ws_rec, ['Item', 'Value'])
        summary_rows = [
            ('Transactions extracted',  recon['txn_count']),
            ('Opening Balance (PDF)',    recon['opening']),
            ('Total Credits (+)',        recon['total_credits']),
            ('Total Debits (−)',         abs(recon['total_debits'])),
            ('Net Movement',             recon['net']),
            ('Computed Closing',         recon['computed_closing']),
            ('Closing Balance (PDF)',    recon['closing_stated']),
            ('Gap',                      recon['summary_gap']),
            ('Status',                   recon['summary_status']),
        ]
        for label, value in summary_rows:
            ws_rec.append([label, value])
            r = ws_rec.max_row
            # Format currency cells
            if isinstance(value, float):
                ws_rec.cell(r, 2).number_format = '₹#,##0.00'
            # Colour the Status row
            if label == 'Status':
                fill = _OK_FILL if '✓' in str(value) else \
                       _ERR_FILL if '✗' in str(value) else _GAP_FILL
                for cell in ws_rec[r]:
                    cell.fill = fill
                ws_rec.cell(r, 2).font = _BOLD

        ws_rec.append([])

        # Row-by-row gap table
        if recon['row_gaps']:
            ws_rec.append([f"  ⚠ {len(recon['row_gaps'])} balance discontinuity(ies) detected"])
            ws_rec[ws_rec.max_row][0].font = Font(bold=True, color='CC0000')
            _hdr_row(ws_rec, [
                'After row #', 'Prev Date', 'Prev Description',
                'Prev Amount', 'Prev Balance', 'Next Date',
                'Next Description', 'Next Amount', 'Next Balance', 'Gap',
            ])
            for g in recon['row_gaps']:
                ws_rec.append([
                    g['after_row'],
                    g['prev_date'], g['prev_desc'], g['prev_amount'], g['prev_balance'],
                    g['curr_date'], g['curr_desc'], g['curr_amount'], g['curr_balance'],
                    g['gap'],
                ])
                r = ws_rec.max_row
                for cell in ws_rec[r]:
                    cell.fill = _GAP_FILL
                for col in [4, 5, 8, 9, 10]:
                    ws_rec.cell(r, col).number_format = '₹#,##0.00'
        else:
            ws_rec.append(['  ✓ No balance discontinuities detected — all rows consecutive'])
            ws_rec[ws_rec.max_row][0].font = Font(color='006400')  # dark green

        ws_rec.append([])
        ws_rec.append(['─' * 80])
        ws_rec.append([])

    ws_rec.column_dimensions['A'].width = 35
    ws_rec.column_dimensions['B'].width = 22
    for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws_rec.column_dimensions[col].width = 20

    wb.save(output_path)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    target = Path(sys.argv[1])
    output_file = None
    if '--output' in sys.argv:
        idx = sys.argv.index('--output')
        if idx + 1 < len(sys.argv):
            output_file = Path(sys.argv[idx + 1])

    configs = load_configs()

    if target.is_dir():
        pdf_files = sorted(target.glob('*.pdf'))
    elif target.suffix.lower() == '.pdf':
        pdf_files = [target]
    else:
        print(f"Error: {target} is not a PDF file or directory")
        sys.exit(1)

    if not pdf_files:
        print(f"No PDF files found in {target}")
        sys.exit(1)

    all_rows   = []
    recon_list = []

    for pdf_path in pdf_files:
        print(f"\nProcessing: {pdf_path.name}")
        rows, opening, closing, _ = process_pdf(pdf_path, configs)
        print(f"  Extracted: {len(rows)} transactions")
        recon = compute_recon(rows, opening, closing, pdf_path.name)
        print(f"  Recon:     {recon['summary_status']}")
        if recon['row_gaps']:
            print(f"  ⚠  {len(recon['row_gaps'])} balance gap(s) detected — see Reconciliation sheet")
        all_rows.extend(rows)
        recon_list.append(recon)

    if not all_rows:
        print("\nNo transactions extracted.")
        sys.exit(0)

    if not output_file:
        stem = target.stem if target.is_file() else target.name
        output_file = target.parent / f"{stem}_transactions.xlsx"

    write_excel(all_rows, recon_list, output_file)

    debits  = [r for r in all_rows if r['amount'] < 0]
    credits = [r for r in all_rows if r['amount'] > 0]
    print(f"\nTotal: {len(all_rows)} transactions → {output_file}")
    print(f"  Debits:  {len(debits):3d} txns   ₹{sum(abs(r['amount']) for r in debits):>14,.2f}")
    print(f"  Credits: {len(credits):3d} txns   ₹{sum(r['amount'] for r in credits):>14,.2f}")


if __name__ == '__main__':
    main()
