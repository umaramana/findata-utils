"""
Excel Column QC & Auto-Correct module.

PDF24 converts broker 1099-B PDFs to Excel with correct column headers,
but data rows can be misaligned in two ways:

1. "Right shift" — Part 1 (Description) consumes extra columns, pushing
   all Part 2 columns (dates, financials) one or more positions right.
2. "Empty cell collapse" — when optional columns (Accrued Market Discount,
   Wash Sale Loss) are empty, values shift left.

Fix: Use the header row as the anchor. For each data row, validate that
the Date Acquired column has a date pattern at the expected position.
If the date is found further right, shift Part 2 values left to re-align
with the header.

Each broker has its own column config — column layouts differ per broker.
"""

import io
import re

import pandas as pd
from openpyxl.styles import PatternFill


# Date pattern: MM/DD/YY, MM/DD/YYYY, or YYYY-MM-DD
_DATE_RE = re.compile(r'^\d{1,2}/\d{1,2}/\d{2,4}$|^\d{4}-\d{1,2}-\d{1,2}$')


# ── Per-broker column configuration ──────────────────────────────────────
# Each broker defines:
#   cost_col_idx:       expected 0-based column index for Cost
#   date_acq_col_idx:   expected 0-based column index for Date Acquired
#   cost_keywords:      keywords to verify/find the Cost column header
#   date_acq_keywords:  keywords to verify/find the Date Acquired header
#   optional_cols:      names of the optional columns in the zone (for logging)

BROKER_CONFIG = {
    'fidelity': {
        # Fidelity 1099-B column order (0-indexed):
        # 0: Action
        # 1: Quantity
        # 2: 1b Date Acquired
        # 3: 1c Date Sold or Disposed
        # 4: 1d Proceeds
        # 5: 1e Cost or Other Basis
        # 6: 1f Accrued Market Discount  ← optional
        # 7: 1g Wash Sale Loss           ← optional
        # 8: Gain/Loss                   ← always present
        'date_acq_col_idx': 2,
        'cost_col_idx': 5,
        'date_acq_keywords': ['date', 'acquired', '1b'],
        'cost_keywords': ['cost', 'basis', '1e'],
        'optional_cols': ['1f Accrued Market Discount', '1g Wash Sale Loss'],
        'gain_loss_col_idx': 8,
        'fed_tax_col_idx': None,
    },
    'charles_schwab': {
        # Schwab 1099-B column order (0-indexed, 9-col layout):
        # 0: Description / CUSIP
        # 1: Strike price
        # 2: Option expiry
        # 3: Date Acquired code (SC/BC for options)
        # 4: Date Acquired (primary row) / Date Sold (secondary row) — paired
        # 5: Proceeds (1d) — sometimes merged with Cost
        # 6: Cost (1e)
        # 7: Accrued Market Discount (1f) / Wash Sale (1g) — merged
        # 8: Realized Gain/Loss
        'date_acq_col_idx': 4,
        'cost_col_idx': 6,
        'date_acq_keywords': ['date', 'acquired', '1b'],
        'cost_keywords': ['cost', 'basis', '1e'],
        'optional_cols': ['1f/1g Accrued/Wash Sale (merged)'],
        'gain_loss_col_idx': 8,
        'fed_tax_col_idx': None,
    },
    'robinhood': {
        # Robinhood 1099-B column order (0-indexed):
        # 0: 1c Date Sold or Disposed
        # 1: Quantity
        # 2: 1d Proceeds
        # 3: 1b Date Acquired
        # 4: 1e Cost or Other Basis
        # 5: 1g Wash Sale Loss           ← optional
        # 6: Gain/Loss                   ← always present
        'date_acq_col_idx': 3,
        'cost_col_idx': 4,
        'date_acq_keywords': ['date', 'acquired', '1b'],
        'cost_keywords': ['cost', 'basis', '1e'],
        'optional_cols': ['1g Wash Sale Loss'],
        'gain_loss_col_idx': 6,
        'fed_tax_col_idx': None,
    },
    'merrill': {
        # Merrill 1099-B column order (0-indexed):
        # 0: 1a. Description of Property
        # 1: (blank)
        # 2: 1b. Date Acquired
        # 3: 1c. Date Sold or Disposed
        # 4: 1d. Proceeds
        # 5: 1e. Cost Basis
        # 6: 1f. Accrued Market Discount  ← optional
        # 7: 1g. Wash Sale Loss           ← optional
        # 8: Gain or Loss                 ← always present
        'date_acq_col_idx': 2,
        'cost_col_idx': 5,
        'date_acq_keywords': ['date', 'acquired', '1b'],
        'cost_keywords': ['cost', 'basis', '1e'],
        'optional_cols': ['1f Accrued Market Discount', '1g Wash Sale Loss'],
        'gain_loss_col_idx': 8,
        'fed_tax_col_idx': None,
    },
    'morgan_stanley': {
        # Morgan Stanley 1099-B column order (0-indexed):
        # 0: Description (Box 1a)
        # 1: Quantity
        # 2: Date Acquired (Box 1b)
        # 3: Date Sold (Box 1c)
        # 4: Proceeds (Box 1d)
        # 5: Cost or Other Basis (Box 1e)
        # 6: Accrued Market Discount (Box 1f)  ← optional
        # 7: Wash Sale Loss Disallowed (Box 1g) ← optional
        # 8: Gain/(Loss)                        ← always present
        # 9: Federal Income Tax Withheld         ← always present
        'date_acq_col_idx': 2,
        'cost_col_idx': 5,
        'date_acq_keywords': ['date', 'acquired', '1b'],
        'cost_keywords': ['cost', 'basis', '1e'],
        'optional_cols': ['Accrued Market Discount', 'Wash Sale Loss Disallowed'],
        'gain_loss_col_idx': 8,
        'fed_tax_col_idx': 9,
    },
    'apex_clearing': {
        # Apex Clearing 1099-B column order (0-indexed):
        # 0: Date Sold
        # 1: Quantity
        # 2: Proceeds (1d)
        # 3: (empty spacer)
        # 4: Date Acquired (1b)
        # 5: Cost or Other Basis (1e)
        # 6: Wash Sale Loss Disallowed (1g)  ← optional (merged accrued/wash)
        # 7: Gain/Loss                       ← always present
        'date_acq_col_idx': 4,
        'cost_col_idx': 5,
        'date_acq_keywords': ['date', 'acquired', '1b'],
        'cost_keywords': ['cost', 'basis', '1e'],
        'optional_cols': ['1f/1g Accrued/Wash Sale (merged)'],
        'gain_loss_col_idx': 7,
        'fed_tax_col_idx': None,
    },
    'jpmorgan': {
        # JP Morgan 1099-B column order (0-indexed):
        # 0: Type prefix / overflow
        # 1: Description / ticker
        # 2: CUSIP / desc overflow
        # 3: CUSIP cont / "Subtotals"
        # 4: Quantity Sold
        # 5: Date Acquired (1b)
        # 6: Date Sold (1c)
        # 7: Proceeds (1d)
        # 8: Cost (1e)
        # 9: Accrued Mkt Discount (1f)  ← optional
        # 10: Wash Sale Loss (1g)       ← optional
        # 11: Gain/Loss
        'date_acq_col_idx': 5,
        'cost_col_idx': 8,
        'date_acq_keywords': ['date', 'acquired', '1b'],
        'cost_keywords': ['cost', 'basis', '1e'],
        'optional_cols': ['Accrued Market Discount', 'Wash Sale Loss'],
        'gain_loss_col_idx': 11,
        'fed_tax_col_idx': None,
    },
}


def _find_header_row(df):
    """Find the header row index in a DataFrame by keyword matching.
    Picks the row with the most keyword matches (within first 10 rows)
    to avoid false positives from form title rows like 'PROCEEDS FROM
    BROKER & BARTER EXCHANGE TRANSACTIONS'.
    """
    keywords = ['proceeds', 'cost', 'basis', 'gain', 'loss', 'action',
                'quantity', 'date', 'acquired', 'sold',
                '1a', '1b', '1c', '1d', '1e', '1f', '1g',
                'wash', 'accrued', 'discount', 'description']

    best_idx = 0
    best_count = 0
    for idx, row in df.head(10).iterrows():
        row_text = ' '.join(str(v).lower() if pd.notna(v) else '' for v in row)
        count = sum(1 for kw in keywords if kw in row_text)
        if count > best_count:
            best_count = count
            best_idx = idx
    return best_idx


def _verify_or_search_col(df, header_idx, expected_idx, keywords):
    """Verify a column at expected_idx matches keywords, or trust config.
    Returns the column index or None.
    """
    col_idx = expected_idx
    if col_idx is not None:
        cell = str(df.iat[header_idx, col_idx]).lower() if col_idx < len(df.columns) else ''
        if not any(kw in cell for kw in keywords):
            # Verify failed — trust config rather than searching.
            # Search was unreliable for paired/multi-row headers (Schwab, Morgan Stanley, JP Morgan).
            # See PARKING_LOT.md for follow-up on removing search entirely.
            col_idx = expected_idx

    return col_idx


def _find_anchor_cols(df, broker_name):
    """Find Date Acquired and Cost column indices from the header row."""
    config = BROKER_CONFIG.get(broker_name)
    if config is None:
        raise ValueError(f"No QC config for broker: {broker_name}")

    header_idx = _find_header_row(df)

    date_acq_col = _verify_or_search_col(
        df, header_idx, config['date_acq_col_idx'], config['date_acq_keywords'])
    cost_col = _verify_or_search_col(
        df, header_idx, config['cost_col_idx'], config['cost_keywords'])

    missing = []
    if date_acq_col is None:
        missing.append('Date Acquired')
    if cost_col is None:
        missing.append('Cost')
    if missing:
        raise ValueError(f"Could not find column(s): {', '.join(missing)}")

    return {'date_acq_col': date_acq_col, 'cost_col': cost_col, 'header_idx': header_idx}


def _is_empty(val):
    """Check if a cell value is empty/NaN."""
    if val is None:
        return True
    if isinstance(val, float) and pd.isna(val):
        return True
    s = str(val).strip()
    return s == '' or s in ('nan', 'NaN', 'None')


def _has_date(val):
    """Check if a cell contains a date pattern (MM/DD/YY), including \\n-separated values."""
    if _is_empty(val):
        return False
    lines = [l.strip() for l in str(val).split('\n') if l.strip()]
    return any(_DATE_RE.match(l) for l in lines) or str(val).strip().upper() == 'VARIOUS'


def _is_numeric(val):
    """Check if a cell value looks like a number (including currency formats)."""
    if _is_empty(val):
        return False
    cleaned = str(val).replace('$', '').replace(',', '').strip()
    if cleaned.startswith('(') and cleaned.endswith(')'):
        cleaned = cleaned[1:-1]
    try:
        float(cleaned)
        return True
    except (ValueError, TypeError):
        return False


def _fix_date_acq_left_shift(df, row_idx, date_acq_col, num_cols):
    """Fix Date Acquired shifted one position left. Returns 1 if fixed, 0 otherwise."""
    if date_acq_col <= 0 or date_acq_col >= num_cols:
        return 0
    if not _is_empty(df.iat[row_idx, date_acq_col]):
        return 0
    check_col = date_acq_col - 1
    if check_col >= 0 and _has_date(df.iat[row_idx, check_col]):
        df.iat[row_idx, date_acq_col] = df.iat[row_idx, check_col]
        df.iat[row_idx, check_col] = None
        return 1
    return 0


def _fix_gain_loss_collapse(df, row_idx, cost_col, gain_loss_col, num_cols):
    """Fix Gain/Loss collapsed left into optional zone. Returns 1 if fixed, 0 otherwise."""
    if gain_loss_col >= num_cols:
        return 0
    if not _is_empty(df.iat[row_idx, gain_loss_col]):
        return 0
    for col in range(gain_loss_col - 1, cost_col, -1):
        if col < num_cols and _is_numeric(df.iat[row_idx, col]):
            df.iat[row_idx, gain_loss_col] = df.iat[row_idx, col]
            df.iat[row_idx, col] = None
            return 1
    return 0


def _fix_fed_tax_shift(df, row_idx, fed_tax_col, gain_loss_col, num_cols):
    """Fix Fed Tax Withheld shifted one position left."""
    if fed_tax_col is None or fed_tax_col >= num_cols:
        return
    if _is_empty(df.iat[row_idx, fed_tax_col]):
        if fed_tax_col - 1 > gain_loss_col and not _is_empty(df.iat[row_idx, fed_tax_col - 1]):
            df.iat[row_idx, fed_tax_col] = df.iat[row_idx, fed_tax_col - 1]
            df.iat[row_idx, fed_tax_col - 1] = None


def _fix_left_shifts(df, broker_name, header_idx, anchors):
    """Pass 2: Fix left-shift / empty-cell-collapse artifacts. Returns number of fixes."""
    config = BROKER_CONFIG.get(broker_name)
    if config is None:
        return 0

    gain_loss_col = config.get('gain_loss_col_idx')
    if gain_loss_col is None:
        return 0

    cost_col = anchors['cost_col']
    date_acq_col = anchors['date_acq_col']
    fed_tax_col = config.get('fed_tax_col_idx')
    num_cols = len(df.columns)
    fixes = 0

    for row_idx in range(header_idx + 1, len(df)):
        if cost_col >= num_cols or not _is_numeric(df.iat[row_idx, cost_col]):
            continue

        fixes += _fix_date_acq_left_shift(df, row_idx, date_acq_col, num_cols)

        gl_fix = _fix_gain_loss_collapse(df, row_idx, cost_col, gain_loss_col, num_cols)
        fixes += gl_fix
        if gl_fix:
            _fix_fed_tax_shift(df, row_idx, fed_tax_col, gain_loss_col, num_cols)

    return fixes


def _fix_right_shift_row(df, row_idx, expected_date_col):
    """Detect and fix a right-shifted row. Returns (offset, True) if fixed, (0, False) otherwise."""
    if expected_date_col >= len(df.columns):
        return 0, False
    if _has_date(df.iat[row_idx, expected_date_col]):
        return 0, False

    actual_date_col = None
    for col in range(expected_date_col + 1, len(df.columns)):
        if _has_date(df.iat[row_idx, col]):
            actual_date_col = col
            break

    if actual_date_col is None:
        return 0, False

    offset = actual_date_col - expected_date_col
    num_cols = len(df.columns)
    for col in range(expected_date_col, num_cols - offset):
        df.iat[row_idx, col] = df.iat[row_idx, col + offset]
    for col in range(num_cols - offset, num_cols):
        df.iat[row_idx, col] = None
    return offset, True


def _collect_review_flags(df, row_idx, sheet_name, cost_col, config, optional_col_names):
    """Check if a corrected row has optional column values that need manual review."""
    num_cols = len(df.columns)
    opt_parts = []
    for col in range(cost_col + 1, cost_col + 1 + len(config['optional_cols'])):
        if col < num_cols:
            val = df.iat[row_idx, col]
            if not _is_empty(val):
                col_name = optional_col_names.get(col, f"col {col}")
                opt_parts.append(f"{col_name} = {val}")
    if opt_parts:
        return f"{sheet_name} Row {row_idx + 1}: {', '.join(opt_parts)}"
    return None


def _process_sheet(df, header_idx, expected_date_col, cost_col, config,
                   optional_col_names, broker_name, anchors, sheet_name):
    """Process one sheet: fix right-shifts (Pass 1) and left-shifts (Pass 2).
    Returns (sheet_fixes, corrected_rows, sheet_details, review_flags).
    """
    sheet_fixes = 0
    sheet_details = []
    corrected_rows = []
    review_flags = []

    for row_idx in range(header_idx + 1, len(df)):
        offset, fixed = _fix_right_shift_row(df, row_idx, expected_date_col)
        if not fixed:
            continue
        sheet_fixes += 1
        corrected_rows.append(row_idx)
        sheet_details.append(
            f"    Row {row_idx + 1}: shifted Part 2 left by {offset} col(s) "
            f"(date was at col {expected_date_col + offset}, expected col {expected_date_col})"
        )
        flag = _collect_review_flags(df, row_idx, sheet_name, cost_col, config, optional_col_names)
        if flag:
            review_flags.append(flag)

    left_fixes = _fix_left_shifts(df, broker_name, header_idx, anchors)
    if left_fixes > 0:
        sheet_fixes += left_fixes
        sheet_details.append(f"    Pass 2: {left_fixes} left-shift correction(s)")

    return sheet_fixes, corrected_rows, sheet_details, review_flags


def _init_qc_context(excel_file, broker_name):
    """Set up QC context: config, anchors, log preamble."""
    config = BROKER_CONFIG.get(broker_name)
    if config is None:
        raise ValueError(f"No QC config for broker: {broker_name}")

    excel_file.seek(0)
    xls = pd.ExcelFile(excel_file)
    first_df = xls.parse(xls.sheet_names[0], header=None)
    anchors = _find_anchor_cols(first_df, broker_name)

    expected_date_col = anchors['date_acq_col']
    cost_col = anchors['cost_col']
    optional_col_names = {cost_col + 1 + i: name for i, name in enumerate(config['optional_cols'])}

    log = [f"Broker: {broker_name}",
           f"Anchor columns: Date Acquired=col {expected_date_col}, Cost=col {cost_col}"]
    if config['optional_cols']:
        log.append(f"Optional zone: {', '.join(config['optional_cols'])}")

    return config, xls, anchors, expected_date_col, cost_col, optional_col_names, log


def _correct_all_sheets(excel_file, xls, config, anchors, expected_date_col,
                        cost_col, optional_col_names, broker_name, log):
    """Run QC corrections across all sheets, write corrected Excel."""
    total_fixes = 0
    any_corrections = False
    review_rows = []

    excel_file.seek(0)
    output = io.BytesIO()
    highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name in xls.sheet_names:
            excel_file.seek(0)
            df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None, dtype=str)
            header_idx = _find_header_row(df)

            sheet_fixes, corrected_rows, sheet_details, flags = _process_sheet(
                df, header_idx, expected_date_col, cost_col, config,
                optional_col_names, broker_name, anchors, sheet_name)
            review_rows.extend(flags)

            if sheet_fixes > 0:
                any_corrections = True
                log.append(f"{sheet_name}: {sheet_fixes} rows corrected")
                log.extend(sheet_details)
            else:
                log.append(f"{sheet_name}: no corrections needed")

            total_fixes += sheet_fixes
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

            if corrected_rows:
                ws = writer.sheets[sheet_name]
                for row_idx in corrected_rows:
                    ws.cell(row=row_idx + 1, column=expected_date_col + 1).fill = highlight_fill

    output.seek(0)
    return output if any_corrections else None, total_fixes, review_rows


def detect_and_correct(excel_file, broker_name):
    """
    Detect and correct column misalignment across all sheets.

    Returns:
        dict with corrected_excel, log, total_fixes, sheets_checked, review_rows.
    """
    config, xls, anchors, expected_date_col, cost_col, optional_col_names, log = \
        _init_qc_context(excel_file, broker_name)

    corrected_excel, total_fixes, review_rows = _correct_all_sheets(
        excel_file, xls, config, anchors, expected_date_col, cost_col,
        optional_col_names, broker_name, log)

    return {
        'corrected_excel': corrected_excel,
        'log': log,
        'total_fixes': total_fixes,
        'sheets_checked': len(xls.sheet_names),
        'review_rows': review_rows,
    }
