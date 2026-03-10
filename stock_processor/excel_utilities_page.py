"""
Excel Utilities — Collate multiple Excel tabs into a single master sheet
with type-aware formatting and an inline reconciliation summary.
"""

import io

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

_NUMBER_FORMATS = {
    'date':     'MM/DD/YYYY',
    'currency': '#,##0.00',
    'integer':  '#,##0',
}


# ── Type helpers ──────────────────────────────────────────────────────────────

def _try_date(val):
    try:
        dt = pd.to_datetime(str(val), dayfirst=False)
        return dt.to_pydatetime() if 1990 <= dt.year <= 2035 else None
    except Exception:
        return None


def _try_number(val):
    try:
        return float(str(val).replace(',', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return None


# ── Collation ─────────────────────────────────────────────────────────────────

def _is_header_row(row):
    non_empty = [v for v in row if pd.notna(v) and str(v).strip() != '']
    if not non_empty:
        return False
    for val in non_empty:
        try:
            float(str(val).replace(',', '').replace('$', ''))
            return False
        except ValueError:
            pass
    return True


def _collate_sheets(uploaded_file, selected_sheets):
    xl = pd.ExcelFile(uploaded_file)
    master_rows, header, max_cols = [], None, 0
    for sheet_name in selected_sheets:
        df = xl.parse(sheet_name, header=None, dtype=str)
        if df.empty:
            continue
        start_row = 0
        if _is_header_row(df.iloc[0]):
            if header is None:
                header = list(df.iloc[0])
            start_row = 1
        data_df = df.iloc[start_row:].dropna(how='all')
        for row in data_df.itertuples(index=False, name=None):
            master_rows.append([sheet_name] + list(row))
            max_cols = max(max_cols, len(row))
    if not master_rows:
        return pd.DataFrame()
    data_cols = header if header else [f'Col{i+1}' for i in range(max_cols)]
    cols = ['Month'] + data_cols[:max_cols]
    padded = [r + [''] * (len(cols) - len(r)) for r in master_rows]
    return pd.DataFrame(padded, columns=cols)


# ── Column type detection & UI ────────────────────────────────────────────────

def _preview_cols(uploaded_file, sheet_name):
    df = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=None, nrows=20, dtype=str)
    if df.empty:
        return ['Month'], {'Month': 'text'}
    start_row = 1 if _is_header_row(df.iloc[0]) else 0
    raw_cols = list(df.iloc[0]) if start_row == 1 else [f'Col{i+1}' for i in range(len(df.columns))]
    sample = df.iloc[start_row:]
    auto_types = {'Month': 'text'}
    for i, col in enumerate(raw_cols):
        vals = [str(v) for v in sample.iloc[:, i] if pd.notna(v) and str(v).strip() not in ('', 'nan')]
        if not vals:
            auto_types[col] = 'text'
        elif sum(1 for v in vals if _try_date(v)) / len(vals) >= 0.8:
            auto_types[col] = 'date'
        elif sum(1 for v in vals if _try_number(v) is not None) / len(vals) >= 0.8:
            auto_types[col] = 'number'
        else:
            auto_types[col] = 'text'
    return ['Month'] + raw_cols, auto_types


def _render_type_selector(auto_types):
    number_cols = [col for col, t in auto_types.items() if t == 'number']
    if not number_cols:
        return {col: ('currency' if t == 'number' else t) for col, t in auto_types.items()}
    with st.expander("Column types — auto-detected, adjust if needed", expanded=True):
        integer_cols = st.multiselect(
            "Mark as Integer (whole numbers — e.g. check numbers):",
            options=number_cols,
            help="Unselected numeric columns are treated as Currency and included in reconciliation totals."
        )
    return {col: ('integer' if col in integer_cols else ('currency' if t == 'number' else t))
            for col, t in auto_types.items()}


# ── Excel output ──────────────────────────────────────────────────────────────

def _build_typed_row(row, col_types_list):
    result = []
    for val, col_type in zip(row, col_types_list):
        if pd.isna(val) or str(val).strip() in ('', 'nan'):
            result.append('')
        elif col_type == 'date':
            result.append(_try_date(val) or str(val))
        elif col_type in ('currency', 'integer'):
            converted = _try_number(val)
            result.append(converted if converted is not None else str(val))
        else:
            result.append(str(val))
    return result


def _col_sum(series):
    """Vectorised sum of a string currency column."""
    return pd.to_numeric(
        series.astype(str).str.replace(',', '', regex=False).str.replace('$', '', regex=False),
        errors='coerce'
    ).fillna(0).sum()


def _append_recon(ws, master_df, col_types):
    """Append reconciliation summary at bottom of Master sheet."""
    currency_cols = [col for col, t in col_types.items() if t == 'currency']
    if not currency_cols:
        return
    ws.append([])
    ws.append(['Reconciliation'])
    ws.append(['Tab', 'Rows'] + currency_cols)
    recon_header_row = ws.max_row
    grand_totals = {col: 0.0 for col in currency_cols}
    grand_rows = 0
    for tab_name, group in master_df.groupby('Month', sort=False):
        row_vals = [tab_name, len(group)]
        for col in currency_cols:
            total = float(_col_sum(group[col]))
            grand_totals[col] += total
            row_vals.append(total)
        ws.append(row_vals)
        grand_rows += len(group)
    ws.append(['Grand Total', grand_rows] + [grand_totals[col] for col in currency_cols])
    for col_idx in range(3, len(currency_cols) + 3):
        for row_idx in range(recon_header_row + 1, ws.max_row + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = '#,##0.00'


def _generate_output(uploaded_file, master_df, col_types):
    """Add Master tab (first position) with inline reconciliation to the original workbook."""
    wb = load_workbook(uploaded_file, data_only=True, keep_links=False)
    for name in ['Master', 'Reconciliation']:
        if name in wb.sheetnames:
            del wb[name]
    ws_m = wb.create_sheet('Master')
    ws_m.append(list(master_df.columns))
    col_types_list = [col_types.get(col, 'text') for col in master_df.columns]
    for _, row in master_df.iterrows():
        ws_m.append(_build_typed_row(row, col_types_list))
    for col_idx, col in enumerate(master_df.columns, start=1):
        fmt = _NUMBER_FORMATS.get(col_types.get(col, 'text'))
        if fmt:
            for (cell,) in ws_m.iter_rows(min_row=2, max_row=ws_m.max_row, min_col=col_idx, max_col=col_idx):
                cell.number_format = fmt
    _append_recon(ws_m, master_df, col_types)
    wb._sheets.remove(ws_m)
    wb._sheets.insert(0, ws_m)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── Page render ───────────────────────────────────────────────────────────────

def _render_tab_collator():
    st.subheader("Collate Excel Tabs → Master")
    uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx", "xls"], key="collator_upload")
    if uploaded_file is None:
        return

    sheet_names = pd.ExcelFile(uploaded_file).sheet_names
    selected = st.multiselect("Select tabs to include:", options=sheet_names, default=sheet_names)
    if not selected:
        st.warning("Select at least one tab.")
        return

    uploaded_file.seek(0)
    _, auto_types = _preview_cols(uploaded_file, selected[0])
    col_types = _render_type_selector(auto_types)

    if st.button("Collate", type="primary"):
        try:
            uploaded_file.seek(0)
            with st.spinner(f"Collating {len(selected)} tab(s)..."):
                master_df = _collate_sheets(uploaded_file, selected)
            if master_df.empty:
                st.error("No data found in selected tabs.")
                return
            uploaded_file.seek(0)
            with st.spinner("Generating output file..."):
                output = _generate_output(uploaded_file, master_df, col_types)
            st.success(f"{len(master_df)} rows from {len(selected)} tab(s) — Master tab added with reconciliation.")
            st.download_button(
                label="Download",
                data=output,
                file_name=uploaded_file.name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )
        except Exception as e:
            st.error(f"Error: {str(e)}")


st.title("Excel Utilities")
st.markdown("---")
_render_tab_collator()
