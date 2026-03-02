"""
Shared utilities for stock transaction processing.
Handles Excel I/O, date parsing, and currency parsing for PDF24-converted files.
"""

import pandas as pd
import numpy as np
import re
import io
from datetime import datetime
from dateutil import parser as date_parser


def parse_date_robust(value):
    """
    Robust date parser handling PDF24 conversion quirks.

    Handles:
    - String dates: "05/02/25", "2024-05-02", "01/11/24"
    - Datetime objects
    - Special values: "Various", "VARIOUS"
    - Excel serial dates
    - Text with trailing characters

    Returns: Standardized date string (MM/DD/YYYY) or original value for special cases
    """
    if pd.isna(value) or value == '' or value is None:
        return ''

    # Handle special values
    str_value = str(value).strip()
    if str_value.upper() == 'VARIOUS':
        return 'Various'

    # Already a datetime object
    if isinstance(value, (datetime, pd.Timestamp)):
        return value.strftime('%m/%d/%Y')

    # Try to parse string dates
    try:
        # Clean up the value
        cleaned = re.sub(r'[\$\s]+$', '', str_value)

        # Try dateutil parser (handles most formats)
        parsed = date_parser.parse(cleaned, dayfirst=False)
        return parsed.strftime('%m/%d/%Y')
    except (ValueError, TypeError):
        pass

    # Try Excel serial date (numeric value)
    try:
        if isinstance(value, (int, float)) and 1 < value < 100000:
            # Excel serial date
            excel_epoch = datetime(1899, 12, 30)
            parsed = excel_epoch + pd.Timedelta(days=int(value))
            return parsed.strftime('%m/%d/%Y')
    except (ValueError, TypeError):
        pass

    # Return original if we couldn't parse
    return str_value


def parse_currency(value):
    """
    Parse currency values from text.

    Handles:
    - "$14,353.09" -> 14353.09
    - "14353.09" -> 14353.09
    - "(100.50)" -> -100.50 (negative in parentheses)
    - "2,184.79 W" -> 2184.79 (wash sale notation)
    - Text with trailing/leading spaces
    - "--" or "..." as zero/empty

    Returns: Float value or None
    """
    if pd.isna(value) or value == '' or value is None:
        return None

    str_value = str(value).strip()

    # Handle empty indicators
    if str_value in ['--', '...', '-', 'nan', 'NaN', 'None']:
        return None

    # Check for negative in parentheses: (100.50) -> -100.50
    paren_match = re.match(r'^\(([\d,.]+)\)$', str_value)
    if paren_match:
        num_str = paren_match.group(1).replace(',', '')
        try:
            return -float(num_str)
        except ValueError:
            return None

    # Remove currency symbols, commas, spaces, and trailing letters (like "W" for wash sale)
    cleaned = re.sub(r'[\$,\s]', '', str_value)
    cleaned = re.sub(r'\s*[A-Za-z]+\s*$', '', cleaned)  # Remove trailing letters

    try:
        return float(cleaned)
    except ValueError:
        return None


def clean_numeric_string(value):
    """
    Remove non-numeric characters except decimal and minus.
    "156.26 W" -> "156.26"
    "$1,234.56" -> "1234.56"
    """
    if pd.isna(value) or value == '' or value is None:
        return ''

    str_value = str(value).strip()

    # Handle empty indicators
    if str_value in ['--', '...', '-', 'nan', 'NaN', 'None']:
        return ''

    # Remove everything except digits, decimal point, and minus sign
    cleaned = re.sub(r'[^\d.\-]', '', str_value)
    return cleaned


def is_date_value(value):
    """
    Check if a value matches common date formats or is the special "VARIOUS" value.
    Used for Schwab processing.
    """
    if pd.isna(value):
        return False

    str_value = str(value).strip()

    # Check for special value "VARIOUS"
    if str_value.upper() == "VARIOUS":
        return True

    # Clean up trailing characters
    cleaned_value = re.sub(r'[\$\s]+$', '', str_value)

    # Common date patterns
    date_patterns = [
        r'^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}$',  # MM/DD/YYYY
        r'^\d{4}[/-]\d{1,2}[/-]\d{1,2}$',     # YYYY-MM-DD
        r'^[A-Za-z]{3,9}\s+\d{1,2}(?:st|nd|rd|th)?(?:,)?\s+\d{2,4}$',  # Month name
        r'^\d{1,2}\s+[A-Za-z]{3,9}(?:,)?\s+\d{2,4}$',  # Day Month Year
        r'^\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}:\d{2}$'   # ISO format
    ]

    for pattern in date_patterns:
        if re.match(pattern, cleaned_value, re.IGNORECASE):
            return True

    return False


def detect_header_row(df, patterns, min_matches=3):
    """
    Find header row by pattern matching.

    Args:
        df: DataFrame to search
        patterns: List of keywords to match (e.g., ['description', 'proceeds'])
        min_matches: Minimum number of pattern matches required

    Returns: Header row index or None
    """
    for idx, row in df.iterrows():
        row_text = ' '.join([str(x).lower() if not pd.isna(x) else '' for x in row.values])
        match_count = sum(1 for pattern in patterns if pattern.lower() in row_text)
        if match_count >= min_matches:
            return idx
    return None


def generate_excel_download(df, filename):
    """
    Generate downloadable Excel file from DataFrame.
    Currency columns are formatted as USD Accounting.
    Returns: BytesIO object for Streamlit download
    """
    from openpyxl.utils import get_column_letter

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Drake_Import')
        ws = writer.sheets['Drake_Import']

        # Apply USD Accounting format to currency columns
        currency_cols = ['Proceeds', 'Cost', 'Wash Sale Loss', 'AMT Cost Basis', 'Accrued Discount']
        for col_idx, col_name in enumerate(df.columns, start=1):
            if col_name in currency_cols:
                col_letter = get_column_letter(col_idx)
                for row in range(2, ws.max_row + 1):  # skip header
                    cell = ws[f'{col_letter}{row}']
                    if cell.value is not None:
                        cell.number_format = '_("$"* #,##0.00_);_("$"* (#,##0.00);_("$"* "-"??_);_(@_)'
    output.seek(0)
    return output


def extract_numeric(val):
    """Return cleaned numeric string, or empty string if not valid.

    Handles: $1,234.56, (100.50) → negative, --, nan, None.
    Shared by Apex Clearing and Charles Schwab for merged accrued/wash columns.
    """
    if pd.isna(val) or val is None:
        return ''
    s = str(val).strip()
    if not s or s.lower() in ('nan', 'none', '--', ''):
        return ''
    cleaned = s.replace('$', '').replace(',', '').strip()
    if cleaned.startswith('(') and cleaned.endswith(')'):
        cleaned = '-' + cleaned[1:-1]
    try:
        float(cleaned)
        return cleaned
    except (ValueError, TypeError):
        return ''


def parse_accrued_wash_sale(val):
    """
    Parse a merged Accrued Market Discount / Wash Sale column.

    Brokers like Apex Clearing and Charles Schwab combine both fields into
    a single column.  For now: treats any non-zero value as Wash Sale Loss.
    Markers "(M)" and "(D)" handling to be added when non-zero test data available.

    Returns dict with 'Accrued Market Discount' and 'Wash Sale Loss'.
    """
    numeric = extract_numeric(val)
    if not numeric:
        return {'Accrued Market Discount': '', 'Wash Sale Loss': ''}
    try:
        if float(numeric) == 0:
            return {'Accrued Market Discount': '', 'Wash Sale Loss': ''}
    except (ValueError, TypeError):
        pass
    return {'Accrued Market Discount': '', 'Wash Sale Loss': numeric}


def clean_dataframe_for_display(df):
    """
    Clean DataFrame for display by replacing NaN/None values with empty strings.
    """
    display_df = df.copy()
    for col in display_df.columns:
        display_df[col] = display_df[col].astype(str)
        display_df[col] = display_df[col].replace(['nan', 'NaN', 'None', 'none'], '')
    return display_df
