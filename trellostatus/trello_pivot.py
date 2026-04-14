# Trello List Pivot — Weekly Saturday routine
# Usage: drop any new Trello board export JSON(s) into this folder, then run:
#        python trello_pivot.py
# Each JSON is processed using its file creation date as the run date.
# Already-recorded dates are skipped automatically.
# Output: trello_list_pivot.xlsx — one file, new date column added per JSON.

import glob
import json
import os
from collections import Counter
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import RichTextProperties
from openpyxl.utils import get_column_letter

OUTPUT_FILE = "trello_list_pivot.xlsx"

# --- Styles ---
BLUE       = "1F4E79"
LIGHT_BLUE = "BDD7EE"
YELLOW     = "FFD966"
WHITE      = "FFFFFF"


def process_json(json_file, wb, ws):
    """Process one JSON file and add a column to the workbook. Returns num_weeks written."""
    run_date = datetime.fromtimestamp(os.path.getctime(json_file)).strftime("%d-%b-%Y")

    with open(json_file, encoding="utf-8") as f:
        data = json.load(f)

    list_names = {lst["id"]: lst["name"] for lst in data["lists"]}

    counter = Counter()
    for card in data["cards"]:
        if not card.get("closed"):
            list_id = card["idList"]
            list_name = list_names.get(list_id, f"Unknown ({list_id})")
            counter[list_name] += 1

    list_order = [lst["name"] for lst in sorted(data["lists"], key=lambda x: x["pos"])]

    # Guard: skip if this date already recorded
    existing_dates = [ws.cell(row=3, column=c).value for c in range(2, ws.max_column + 1)]
    if run_date in existing_dates:
        print(f"  Skipping {run_date} — already recorded.")
        return

    new_col = ws.max_column + 1

    # Build name -> row map from existing col A
    name_to_row = {}
    total_row = None
    for r in range(4, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val == "TOTAL":
            total_row = r
        elif val is not None:
            name_to_row[val] = r

    # If no TOTAL row found, create one at the end
    if total_row is None:
        total_row = ws.max_row + 1
        ws.cell(row=total_row, column=1, value="TOTAL")

    # Append any new lists before TOTAL
    for name in list_order:
        if name not in name_to_row:
            ws.insert_rows(total_row)
            new_row = total_row
            total_row += 1
            ws.cell(row=new_row, column=1, value=name)
            name_to_row[name] = new_row
            print(f"  [new list] {name}")

    # Header
    hdr = ws.cell(row=3, column=new_col, value=run_date)
    hdr.font = Font(bold=True, color=WHITE)
    hdr.fill = PatternFill("solid", fgColor=BLUE)
    hdr.alignment = Alignment(horizontal="center")

    # Data rows — match by name
    for i, name in enumerate(list_order):
        row = name_to_row[name]
        count = counter.get(name, 0)
        name_cell  = ws.cell(row=row, column=1)
        count_cell = ws.cell(row=row, column=new_col, value=count)
        if i % 2 == 0:
            name_cell.fill  = PatternFill("solid", fgColor=LIGHT_BLUE)
            count_cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        count_cell.alignment = Alignment(horizontal="center")

    # Total row
    total_name  = ws.cell(row=total_row, column=1)
    total_count = ws.cell(row=total_row, column=new_col, value=sum(counter.values()))
    for cell in [total_name, total_count]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=YELLOW)
    total_count.alignment = Alignment(horizontal="center")

    # Column width
    ws.column_dimensions[get_column_letter(new_col)].width = 14

    # Re-merge title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=new_col)

    # Rebuild chart
    ws._charts = []
    num_weeks = new_col - 1
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Cards per List — Weekly"
    chart.x_axis.title = "List"
    chart.y_axis.title = "Card Count"
    chart.style = 10
    chart.x_axis.axPos = "b"
    chart.y_axis.axPos = "l"
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    chart.dLbls.showSerName = False
    chart.dLbls.showCatName = False
    chart.x_axis.txPr = RichText(bodyPr=RichTextProperties(rot=-5400000))
    data_ref = Reference(ws, min_col=2, max_col=new_col, min_row=3, max_row=total_row - 1)
    cats_ref = Reference(ws, min_col=1, min_row=4, max_row=total_row - 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width  = 30
    chart.height = 14
    ws.add_chart(chart, f"{get_column_letter(new_col + 2)}1")

    print(f"  {run_date} — {sum(counter.values())} cards across {len(list_order)} lists (week {num_weeks})")
    for name in list_order:
        print(f"    {name:<40} {counter.get(name, 0):>5}")
    print(f"    {'TOTAL':<40} {sum(counter.values()):>5}")


# --- Sort all JSONs by creation date ---
json_files = sorted(glob.glob("*.json"), key=lambda f: os.path.getctime(f))
if not json_files:
    print("No JSON files found.")
    exit(1)

print(f"Found {len(json_files)} JSON file(s): {', '.join(json_files)}")
print()

# --- Load or create workbook using the first JSON ---
if os.path.exists(OUTPUT_FILE):
    wb = load_workbook(OUTPUT_FILE)
    ws = wb["Weekly Status"]
else:
    # Bootstrap workbook from the first JSON
    first_file = json_files[0]
    with open(first_file, encoding="utf-8") as f:
        first_data = json.load(f)
    first_list_order = [lst["name"] for lst in sorted(first_data["lists"], key=lambda x: x["pos"])]

    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Status"

    ws["A1"].value = "Rasrich Tax LLC - 2026 Weekly Status Updates"
    ws["A1"].font = Font(bold=True, size=14, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 6

    cell = ws.cell(row=3, column=1, value="List")
    cell.font = Font(bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=BLUE)

    for i, name in enumerate(first_list_order):
        ws.cell(row=4 + i, column=1, value=name)
    ws.cell(row=4 + len(first_list_order), column=1, value="TOTAL")
    ws.column_dimensions["A"].width = 42

    # Seed max_column so process_json sees col 1 as the last col (new_col will be 2)
    ws.cell(row=3, column=1)  # ensure col A exists as max

for json_file in json_files:
    print(f"Processing: {json_file}")
    process_json(json_file, wb, ws)
    print()

wb.save(OUTPUT_FILE)
print(f"Saved {OUTPUT_FILE}")
